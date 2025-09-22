# -*- coding: utf-8 -*-
"""
Объединенный анализ цен: Cash, F17, Zakup, VIP

Содержит все четыре типа анализа:
- Cash: анализ наличных цен с таргетом на 2-ю цену рынка
- F17: анализ F17 цен с андеркотом p1 на 0.50
- Zakup: анализ закупочных цен
- VIP: анализ VIP цен

Запуск: streamlit run unified_analysis.py
"""

import math
import re
import zipfile
from io import BytesIO
from typing import List, Optional, Tuple, Dict
from decimal import Decimal, ROUND_HALF_UP

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule, CellIsRule
import pandas as pd
import streamlit as st


# ========= Общие параметры =========
STEP = 0.1
MIN_SUPPLIERS = 2                 # минимум валидных цен рынка
LIFT_GAP = 0.5                    # (не используется для таргета p2, но оставлено для совместимости)
DOS_HIGH = 45                     # DoS > 45 → цель #1 (только для отчётных метрик)
DOS_MID  = 20                     # 20 ≤ DoS ≤ 45 → цель #2 (метрика)

GAP_THRESHOLD = 0.40              # анти-выбросы: разрыв ≥40% → отбрасываем край
MARKET_ABOVE_OLD_CAP_PCT = 0.15   # потолок к Zakup_old: +15%

# Cash параметры
CASH_MIN_MARGIN_PCT = 0.015       # общий нижний порог: Cash ≥ Cost×1.015
FIRST_PLACE_MARGIN_PCT = 0.12     # (исторический параметр; не задействован в P2-target)

# F17 параметры
MIN_ABS_MARGIN = 0.50     # абсолютная минимальная наценка к Cost
PRICE_TICK   = 0.01

UNDERCUT_MIN = PRICE_TICK # минимум, на сколько ниже p1 должны быть (0.01)
FALLBACK_COST_COL_1B = None # дефолтная колонка Cost (1-based), если не нашли по заголовку
FALLBACK_ZAKUP_COL_1B = 25 # дефолтная колонка Zakup (1-based), если не нашли по заголовку
FALLBACK_VIP_COL_1B = 36 # дефолтная колонка VIP (1-based), если не нашли по заголовку

# --- F17 alignment with base version ---
ROUND_TO_HALF = True            # floor к шагу 0.50
ROUND_ENDING_99 = True          # .99 только когда итог ровно целое
GRANDFATHER_MIN_MARGIN = 0.00   # разрешаем держать старую, даже если < Cost+0.50
GAP_TIGHTEN_ENABLED = True
GAP_TIGHTEN_TRIGGER = 1.00      # если gap с p1 ≥ 1.00 → подтяжка (разрешаем upmove)
SKIP_OLD_WHEN_MARKET_INFEASIBLE = True  # рынок есть, но infeasible → не откатываемся на старую

# ========= Общие утилиты =========
TRUE_TOK = {"true","да","y","1","ж","желтая","yellow"}
FALSE_TOK = {"false","нет","n","0","","неж","not yellow","notyellow"}

def norm(s: str) -> str:
    s = str(s)
    s = s.lower().replace("ё","е")
    s = re.sub(r"[·–—-−]+","-", s)
    s = s.replace("…","...")
    s = re.sub(r"\s+"," ", s)
    return s.strip()

def is_true_like(v) -> bool:
    if isinstance(v,bool): return v
    if v is None: return False
    if isinstance(v,(int,float)) and pd.notna(v):
        try: return float(v)==1.0
        except: return False
    if isinstance(v,str):
        s = norm(v)
        if s in TRUE_TOK: return True
        if s in FALSE_TOK: return False
        if s.startswith("ж") and "желт" in s: return True
    return False

def parse_float(x) -> Optional[float]:
    if x is None: return None
    if isinstance(x,(int,float)):
        try: return float(x)
        except: return None
    s = re.sub(r"\s+","", str(x).strip()).replace(",", ".")
    if s in {"-","—","–",""}: return None
    try: return float(s)
    except: return None

def parse_price(x) -> Optional[float]:
    if x is None: return None
    if isinstance(x,str):
        s = x.split("/")[0]
        s = re.sub(r"\s+","", s).replace(",", ".")
        if s in {"-","—","–",""}: return None
        try: return float(s)
        except: return None
    if isinstance(x,(int,float)):
        try: return float(x)
        except: return None
    return None

def floor1(x: Optional[float]) -> Optional[float]:
    return math.floor(x*10)/10.0 if x is not None else None

def ceil1(x: Optional[float]) -> Optional[float]:
    return math.ceil(x * 10) / 10.0 if x is not None else None

def round_half_down_0_01(x: Optional[float]) -> Optional[float]:
    if x is None: return None
    d = Decimal(str(x)) - Decimal("0.0005")
    return float(d.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))

# ========= F17 helpers (динамический undercut, округления, soft-cap) =========
from decimal import ROUND_FLOOR

def calc_undercut_step(p_ref: Optional[float]) -> float:
    """
    Базовая F17-логика: если p1 заканчивается на ...9 (последняя цифра центов), шаг = 0.49, иначе 0.50.
    """
    if p_ref is None:
        return 0.50
    d = Decimal(str(p_ref)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    s = f"{d:.2f}"
    return 0.49 if s.endswith("9") else 0.50

def round_to_0_50(x: Optional[float]) -> Optional[float]:
    if x is None: 
        return None
    d = Decimal(str(x))
    # floor к шагу 0.50
    y = (d * 2).to_integral_value(rounding=ROUND_FLOOR) / Decimal(2)
    return float(Decimal(y).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))

def apply_ending_99_if_integer(x: Optional[float]) -> Optional[float]:
    """Если после округления получилось целое — делаем .99"""
    if x is None:
        return None
    d = Decimal(str(x)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    if d == d.to_integral_value():
        d = d - Decimal("0.01")
    return float(d)

def target_by_market_softcap(p_ref: Optional[float], cost: Optional[float], f17_old: Optional[float]) -> Optional[float]:
    if p_ref is None or cost is None:
        return None

    step = calc_undercut_step(p_ref)     # обычно 0.50
    min_allowed = cost + 0.50
    target_max_50 = p_ref - step         # хотим быть ≤ этого

    # если с 0.50 в принципе невыполнимо — сразу пробуем 0.49
    if min_allowed > target_max_50 + 1e-9:
        cand49 = float(Decimal(str(p_ref)) - Decimal("0.49"))
        cand49 = float(Decimal(str(cand49)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
        if cand49 >= min_allowed - 1e-9:
            # cap только на повышения
            if f17_old is not None and cand49 > f17_old:
                cand49 = min(cand49, f17_old, cost + 3.00)
            return cand49
        return None

    # базовый кандидат под 0.50
    cand = max(min_allowed, target_max_50)

    # cap только на повышения (для снижений кап НЕ применяем)
    if f17_old is not None and cand > f17_old:
        cand = min(cand, f17_old, cost + 3.00)

    # округление в «режиме 0.50»
    if ROUND_TO_HALF:
        cand = round_to_0_50(cand)
    if ROUND_ENDING_99:
        cand = apply_ending_99_if_integer(cand)

    # если после floor к 0.50 провалились ниже минимума — пробуем «мягкий» андеркат 0.49 (даёт .99)
    if cand < min_allowed - 1e-9:
        cand49 = float(Decimal(str(p_ref)) - Decimal("0.49"))
        cand49 = float(Decimal(str(cand49)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
        if cand49 >= min_allowed - 1e-9:
            if f17_old is not None and cand49 > f17_old:
                cand49 = min(cand49, f17_old, cost + 3.00)
            return cand49
        # крайний случай — отдать минимум (в центах), если 0.49 тоже не спасает
        return float(Decimal(str(min_allowed)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))

    # подстраховка: держим под p1-step (для 0.50-ветки)
    if cand > target_max_50 + 1e-9:
        cand = target_max_50
        if ROUND_TO_HALF:
            cand = round_to_0_50(cand)
        if ROUND_ENDING_99:
            cand = apply_ending_99_if_integer(cand)
        if cand < min_allowed - 1e-9:
            # fallback к 0.49, если после подрезки опять сломали минимум
            cand49 = float(Decimal(str(p_ref)) - Decimal("0.49"))
            cand49 = float(Decimal(str(cand49)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
            if cand49 >= min_allowed - 1e-9:
                if f17_old is not None and cand49 > f17_old:
                    cand49 = min(cand49, f17_old, cost + 3.00)
                return cand49
            return None

    # не повышаем против валидной старой
    if f17_old is not None and cand > f17_old + 1e-9:
        cand = f17_old

    return float(Decimal(str(cand)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))

def is_intlike(val) -> bool:
    if val is None: return False
    if isinstance(val,int): return True
    if isinstance(val,float): return float(val).is_integer()
    s = str(val).strip()
    if s.endswith(".0"): s = s[:-2]
    return s.isdigit()

def looks_like_name(s) -> bool:
    if not isinstance(s,str): return False
    if s.strip() in {"+","-","=",""}: return False
    return bool(re.search(r"[A-Za-zА-Яа-я]", s)) and len(s.strip())>=3

def extract_name_row(data: List[List], product_row: int, name_col: int=4) -> Optional[str]:
    cand = data[product_row][name_col] if len(data[product_row])>name_col else None
    if looks_like_name(cand): return str(cand).strip()
    for j in range(name_col-2, name_col+3):
        if j<0: continue
        c = data[product_row][j] if len(data[product_row])>j else None
        if looks_like_name(c): return str(c).strip()
    return None

def guess_brand(name: Optional[str]) -> Optional[str]:
    if not name: return None
    s = str(name)
    s = re.sub(r"\b(edp|edt|edc|deo|man|men|woman|lady|tester)\b"," ", s, flags=re.IGNORECASE)
    s = re.sub(r"\b\d+(\.\d+)?\s*ml\b"," ", s, flags=re.IGNORECASE)
    s = re.sub(r"\s+"," ", s).strip()
    token = s.split(" ")[0]
    mapping = {
        "pr.":"P.R.","p.r.":"P.R.","pr":"P.R.",
        "armani":"Armani","emporio":"Armani",
        "dkny":"DKNY","kenzo":"Kenzo","lanvin":"Lanvin",
        "lacoste":"Lacoste","givenchy":"Givenchy","trussardi":"Trussardi",
        "cacharel":"Cacharel","versace":"Versace","gucci":"Gucci",
        "boss":"Hugo Boss","hugo":"Hugo Boss",
        "banderas":"A.Banderas","a.banderas":"A.Banderas","a.banderas.":"A.Banderas",
    }
    key = norm(token)
    return mapping.get(key, token)

def load_workbook_safe(file, data_only=True):
    data = file.read() if hasattr(file,"read") else open(file,"rb").read()
    bio = BytesIO(data)
    if not zipfile.is_zipfile(bio):
        raise ValueError("Файл не .xlsx (zip) — возможно .xls или повреждён.")
    bio.seek(0)
    try:
        return openpyxl.load_workbook(bio, data_only=data_only)
    except KeyError as e:
        if "xl/sharedStrings.xml" in str(e):
            raise ValueError("Нет xl/sharedStrings.xml. Пересохраните как нормальный .xlsx.")
        raise

# ========= Рынок: матчинг поставщиков =========
_MARKET_BASE = [
    "шанталь спец","шанталь","shantal","Шанталь С...","шанталь с...",
    "борики братья - гульден 14","борики братья — гульден 14","борики братья","гульден 14",
    "владимир","владимир спец",
    "дима американец",
    "лужники ирина (пав.1.6)","лужники ирина","пав.1.6",
    "настя марина","настя мар","натс",
    "парк",
    "Саша Черн...","тихом","тоня","сергей",
    "альянс групп","альянс-групп","альянс груп","alyans grupp","alyans group","alliance group",
    "белка","belka",
    "андрей","andrey","andrei",
    "лужники и",
    "авангард","avangard",
]
_EXCL = ["владимир марк","vladimir mark","парк марк","park mark",
         "шанталь марк","шанталь марка","шанталь м","shantal mark","shantal m"]
_EXCL_PAT = [
    r"^\s*владимир\s*\.\.\.\s*$",
    r"(?:^|[\s(\[{])владимир\s*\.{3}(?:$|[\s)\]}:,-])",
    r"(?:^|[\s(\[{])шанталь\s*\.{1,3}(?:$|[\s)\]}:,-])",
    r"(?:^|[\s(\[{])шанталь\s*марк\w*(?:$|[\s)\]}:,-])",
    r"(?:^|[\s(\[{])shantal\s*\.{0,3}(?:$|[\s)\]}:,-])",
    r"(?:^|[\s(\[{])shantal\s*mark\w*(?:$|[\s)\]}:,-])",
]

def _normalize_name_for_match(s: str)->str:
    ns = norm(s)
    ns = re.sub(r"[^\w\s().:-]"," ", ns)
    ns = re.sub(r"\s+"," ", ns).strip()
    return ns

def _word_boundary_contains(haystack: str, needle: str)->bool:
    pat = r"(?:^|[\s().:-])"+re.escape(needle)+r"(?:$|[\s().:-])"
    return re.search(pat, haystack, flags=re.IGNORECASE) is not None

# --- Константы для Parel ---
PAREL_MARK_PATTERNS = [
    r"\bпарел[ьи]\s*ма",
    r"\bпарел[ьи]\s*марк",
    r"\bparel+\s*ma",
    r"\bparel+\s*mark",
    r"\bparell+\s*ma",
    r"\bparell+\s*mark"
]

def _is_parel_mark(ns: str) -> bool:
    return any(re.search(p, ns, flags=re.IGNORECASE) for p in PAREL_MARK_PATTERNS)

def _is_bare_parel(ns: str) -> bool:
    if re.search(r"\bпарел[ьи]\b(?!\s*ма)", ns, flags=re.IGNORECASE):
        return True
    if re.search(r"\bparel{1,2}\b(?!\s*ma)", ns, flags=re.IGNORECASE):
        return True
    return False

def is_ref_supplier(s: str) -> bool:
    """
    Референс для pmin берём ТОЛЬКО из:
    - 'парк марк' (и 'park mark')
    - 'шамбор ндс' / 'шамбор без ндс'
    - 'парель марк' (в т.ч. 'Парель Ма...', 'parel mark', 'parell mark')
    """
    if not isinstance(s, str):
        return False
    ns = _normalize_name_for_match(s)

    # "голую" Парель нигде не используем
    if _is_bare_parel(ns):
        return False

    if "парк марк" in ns or "park mark" in ns:
        return True
    if "шамбор" in ns and ("ндс" in ns or "без нд" in ns):
        return True
    if _is_parel_mark(ns):
        return True
    return False

def is_market_supplier(s: str)->bool:
    if not isinstance(s,str) or not s.strip(): return False
    ns = _normalize_name_for_match(s)
    for bad in _EXCL:
        if _word_boundary_contains(ns, _normalize_name_for_match(bad)): return False
    for pat in _EXCL_PAT:
        if re.search(pat, ns, flags=re.IGNORECASE): return False
    tokens = [t for t in re.split(r"\s+", ns) if t]
    def has_prefix(pref: str)->bool:
        pref = pref.lower()
        return any(t.startswith(pref) for t in tokens)
    if (has_prefix("лужник") or has_prefix("luzh")) and (has_prefix("и") or has_prefix("ир") or has_prefix("irin") or has_prefix("i")):
        return True
    if has_prefix("альянс") or has_prefix("alyans"):
        if has_prefix("груп") or has_prefix("гр") or has_prefix("group") or has_prefix("grp") or has_prefix("gr"):
            return True
    for key in _MARKET_BASE:
        if _word_boundary_contains(ns, _normalize_name_for_match(key)): return True
    if re.fullmatch(r"\s*владимир\s*", ns, flags=re.IGNORECASE): return True
    return False

# ========= Флаги Excel =========

def _hex_to_rgb(argb: str):
    if not argb: return None
    s = re.sub(r"[^0-9A-Fa-f]","", str(argb))
    s = s[-6:] if len(s)>=6 else s
    if len(s)!=6: return None
    try: return (int(s[0:2],16), int(s[2:4],16), int(s[4:6],16))
    except: return None

def _is_yellowish_rgb(rgb)->bool:
    if rgb is None: return False
    r,g,b = rgb
    return (r>=200 and g>=190 and b<=170) or (r>=245 and g>=240 and b>=200)

def _color_to_hex_upper(col)->str:
    if not col: return ""
    try:
        rv = getattr(col,"rgb",None)
        if isinstance(rv,str): return rv.upper()
        if rv is not None:
            inner = getattr(rv,"rgb",None)
            if isinstance(inner,str): return inner.upper()
            return str(rv).upper()
    except: pass
    try:
        v = getattr(col,"value",None)
        if isinstance(v,str): return v.upper()
    except: pass
    return ""

def _cell_is_yellow(cell)->bool:
    fill = getattr(cell,"fill",None)
    if not fill:
        return False
    if str(getattr(fill,"fill_type","")).lower()!="solid":
        return False

    for attr in ("fgColor","start_color","end_color"):
        col = getattr(fill, attr, None)
        if not col:
            continue
        raw = _color_to_hex_upper(col)
        if raw.endswith("FFFF00") or raw in {"00FFFF00","FFFFFF00","FFFFFF99","FFFFFBF0"}:
            return True
        idx = getattr(col, "indexed", None)
        if isinstance(idx, int) and idx in (5,6,13,43,44):
            return True
        rgb = _hex_to_rgb(raw)
        if _is_yellowish_rgb(rgb):
            return True

    return False

def _cell_is_strike(cell)->bool:
    f = getattr(cell,"font",None)
    return bool(getattr(f,"strike",False) or getattr(f,"strikethrough",False)) if f else False

def add_yellow_flags_to_df(file)->pd.DataFrame:
    wb = load_workbook_safe(file, data_only=True)
    ws = wb.active
    max_col = ws.max_column
    header = [cell.value for cell in ws[1][:max_col]]
    yellow_header = [f"Yellow Flag Col{col}" for col in range(1,max_col+1)]
    strike_header = [f"Strike Flag Col{col}" for col in range(1,max_col+1)]
    rows = [header + yellow_header + strike_header]
    for row in ws.iter_rows(min_row=2, max_col=max_col):
        values = [c.value for c in row]
        y_flags = [_cell_is_yellow(c) for c in row]
        s_flags = [_cell_is_strike(c) for c in row]
        rows.append(values + y_flags + s_flags)
    return pd.DataFrame(rows)

def detect_flag_base(data: List[List])->Optional[int]:
    if not data: return None
    for j,v in enumerate(data[0]):
        if isinstance(v,str) and v.strip().startswith("Yellow Flag Col"): return j
    return None

def detect_strike_base(data: List[List])->Optional[int]:
    if not data: return None
    for j,v in enumerate(data[0]):
        if isinstance(v,str) and v.strip().startswith("Strike Flag Col"): return j
    return None

def flag_at(data: List[List], row_idx: int, price_col_idx0: int, flag_base: int)->bool:
    idx = flag_base + price_col_idx0
    val = data[row_idx][idx] if (0<=row_idx<len(data) and len(data[row_idx])>idx) else None
    return is_true_like(val)

# ========= Остатки =========
# Разрешаем 12,0 и 12.0, но всё равно фильтруем только числа/дефисы
STOCK_PATTERN = re.compile(r"^\s*([0-9]+(?:[.,]\d+)?|[-])(\s*/\s*([0-9]+(?:[.,]\d+)?|[-]))+\s*$")

# НОВОЕ: формат «1 771 шт» / «1771 pcs»
PLAIN_STOCK_PATTERN = re.compile(r"^\s*([0-9][0-9\s.,]*)\s*(?:шт|pcs)\b", re.IGNORECASE)

def parse_stock_plain(val) -> Optional[int]:
    if not isinstance(val, str):
        return None
    m = PLAIN_STOCK_PATTERN.match(val.strip())
    if not m:
        return None
    digits = re.sub(r"[^0-9]", "", m.group(1))
    return int(digits) if digits else None

# строгий plain-формат: "1 771 шт" или "1771 pcs" — ЮНИТ обязателен
PLAIN_STOCK_STRICT = re.compile(r"^\s*([0-9][0-9\s.,]*)\s*(?:шт|pcs)\b", re.IGNORECASE)

def parse_stock_plain_strict(val) -> Optional[int]:
    if not isinstance(val, str):
        return None
    s = val.replace("\xa0", " ").strip()
    m = PLAIN_STOCK_STRICT.match(s)
    if not m:
        return None
    digits = re.sub(r"[^0-9]", "", m.group(1))
    return int(digits) if digits else None

def parse_multi_stock_tokens(val) -> Tuple[Optional[int], Optional[int]]:
    """
    В одной ячейке может быть несколько списков A/B/... разделённых пробелами.
    Складываем реальный остаток по всем токенам; виртуал — суммой последних сегментов.
    """
    if not isinstance(val, str):
        return None, None
    s = val.replace("\xa0", " ")
    total, virt, found = 0, 0, False
    for tok in s.split():
        if STOCK_PATTERN.match(tok):
            rv = parse_stock_value(tok) or 0
            vv = parse_virtual_stock_value(tok) or 0
            total += rv
            virt  += vv
            found = True
    return (total, virt) if found else (None, None)

def parse_stock_value(val) -> Optional[int]:
    """
    Суммируем все сегменты КРОМЕ последнего (последний — виртуальный).
    Если все сегменты '-' → вернуть 0 (а не None).
    """
    if val is None:
        return None
    s = str(val).strip()
    if not s or "/" not in s:
        return None
    parts = [p.strip() for p in s.split("/")]
    if len(parts) < 2:
        return None

    real = parts[:-1]  # последний — виртуал
    total = 0
    had_digit = False
    for p in real:
        if p in {"", "-", "—", "–"}:
            continue
        try:
            # поддержим и "28,0"
            v = float(p.replace(",", "."))
            total += int(v)
            had_digit = True
        except:
            # если мусор — игнорируем сегмент
            pass

    if had_digit:
        return total
    else:
        # все сегменты были '-' → это реальный 0
        return 0

def parse_virtual_stock_value(val) -> Optional[int]:
    if val is None:
        return None
    s = str(val).strip()
    if not s or "/" not in s:
        return None
    last = s.split("/")[-1].strip()
    if last in {"", "-", "—", "–"}:
        return 0
    try:
        return int(float(last.replace(",", ".")))
    except:
        return None

def read_stock_robust(data: List[List], pr: int, y_base: int, stock_col_guess: Optional[int]) -> Tuple[Optional[int], Optional[int], Optional[str]]:
    """
    Ищем остаток только в «остаточных» колонках:
      1) сначала stock_col_guess (если найден),
      2) затем любые колонки, чьи заголовки (верх или строка товара) похожи на «остаток/шт/stock»,
      3) если и так ничего — сканируем ВСЕ колонки, но принимаем только:
         - списки A/B/... (STOCK_PATTERN),
         - либо «plain» со строгим юнитом (шт|pcs).
    Приоритет кандидатов: plain (юнит) > список; затем — по величине total.
    """
    def _candidate_cols():
        cols = []
        if stock_col_guess is not None and stock_col_guess >= 0:
            cols.append(stock_col_guess)
        for j in range(0, y_base):
            if j in cols:
                continue
            top = data[0][j] if (len(data) > 0 and len(data[0]) > j) else None
            hdr = data[pr][j] if (pr < len(data) and len(data[pr]) > j) else None
            if _is_stockish_header(top) or _is_stockish_header(hdr):
                cols.append(j)
        return cols

    cand_cols = _candidate_cols()
    scan_all  = (len(cand_cols) == 0)

    candidates = []  # (prio, total, rr, jj, raw, virt), prio: 2=plain, 1=list

    def _check_cell(rr, jj, v):
        if not isinstance(v, str):
            return
        s = v.replace("\xa0", " ")

        # несколько списков в одной ячейке
        tot_m, virt_m = parse_multi_stock_tokens(s)
        if tot_m is not None:
            candidates.append((1, tot_m, rr, jj, v, virt_m or 0))
            return

        # строгий plain с юнитом
        p = parse_stock_plain_strict(s)
        if p is not None:
            candidates.append((2, p, rr, jj, v, 0))

    for r_off in (0, 1, 2):
        rr = pr + r_off
        if rr >= len(data):
            break
        cols_iter = range(0, y_base) if scan_all else cand_cols
        for jj in cols_iter:
            if len(data[rr]) <= jj:
                continue
            _check_cell(rr, jj, data[rr][jj])

    if not candidates:
        return None, None, None

    # plain важнее списка; затем максимальный total
    candidates.sort(key=lambda t: (t[0], t[1]), reverse=True)
    _prio, total, rr, jj, raw, virt = candidates[0]
    return total, virt, raw

# ========= Поиск колонок =========
# ========= Поиск колонок COST / CASH =========
COST_HEADER_TOKENS  = {"cost", "себестоим", "себест", "с/стоим", "себес"}
CASH_HEADER_TOKENS  = {"cash", "кэш", "нал", "наличный", "наличные"}
F17_HEADER_TOKENS  = {"f17", "f 17", "f-17", "f17 €", "f17, €", "ф17", "ф 17", "ф-17"}
ZAKUP_HEADER_TOKENS = {"закупочная", "закуп", "закупка"}
VIP_HEADER_TOKENS = {"vip", "vip цена", "vip price"}
STOCK_HEADER_TOKENS = {"остаток", "остатки", "stock"}
STOCK_HEADER_TOKENS_EXT = {"остаток", "остатки", "stock", "шт", "pcs"}

def _is_stockish_header(text: str) -> bool:
    return isinstance(text, str) and any(tok in norm(text) for tok in STOCK_HEADER_TOKENS_EXT)
FALLBACK_VIP_COL_1BASED  = 36

# ========= ЗАМОРОЗКА КОНСТАНТ =========
# Замораживаем критические константы, чтобы изменения глобалок не влияли на Cash/VIP

# Cash константы (заморожены)
_CASH_MIN_MARGIN_PCT_FROZEN = CASH_MIN_MARGIN_PCT
_MIN_SUPPLIERS_FROZEN = MIN_SUPPLIERS
_GAP_THRESHOLD_FROZEN = GAP_THRESHOLD

# VIP константы (заморожены)
_VIP_FALLBACK_COL_FROZEN = FALLBACK_VIP_COL_1BASED

# --- ЧИСТЫЙ COST: только RUB, без OLD/EUR, только жёлтые значения ---

COST_OK_TOKENS = {"cost", "себестоимость", "кост"}
COST_BAD_KEYWORDS = {
    "eur","€","usd","$","доллар","евро","cny",
    "old","стар","архив","prev","предыд","hist","history"
}

def _is_clean_cost_header(text: str) -> bool:
    if not isinstance(text, str): 
        return False
    s = norm(text)
    if not any(tok in s for tok in COST_OK_TOKENS):
        return False
    if any(bad in s for bad in COST_BAD_KEYWORDS):
        return False
    # отсечь явные «min», «порог» и т.п.
    if any(k in s for k in ("min", "мин", "порог", "target", "1.5", "1,5", "min15")):
        return False
    return True

def _gather_headers_pairwise(data: List[List], pr: int, j: int) -> str:
    top = data[0][j] if (len(data)>0 and len(data[0])>j) else None
    hdr = data[pr][j] if (pr < len(data) and len(data[pr])>j) else None
    parts = []
    if top: parts.append(str(top))
    if hdr and hdr != top: parts.append(str(hdr))
    return " / ".join(parts) if parts else ""

def find_cost_cols_clean(data: List[List], pr: int, y_base: int) -> List[int]:
    cols = []
    for j in range(0, y_base):
        header_text = _gather_headers_pairwise(data, pr, j)
        if _is_clean_cost_header(header_text):
            cols.append(j)
    # приоритезируем «идеальные» заголовки
    def _score(j):
        h = norm(_gather_headers_pairwise(data, pr, j))
        score = 0
        if h in {"cost","себестоимость"}: score += 100
        if "руб" in h or "₽" in h or " rub" in h: score += 20
        return score
    cols.sort(key=_score, reverse=True)
    return cols

def read_cost_value_strict(
    data: List[List], pr: int, y_base: int, s_base: int
) -> Tuple[Optional[float], Optional[int], Optional[str], Optional[int]]:
    """
    Возвращает: (cost_val, cost_col, cost_src, cost_row_idx)
    Берём ТОЛЬКО жёлтое, не зачёркнутое значение.
    Ищем в строках pr+1, потом pr+2. Если нет — None.
    """
    cols = find_cost_cols_clean(data, pr, y_base)
    if not cols:
        return None, None, None, None

    for r_off in (1, 2):
        rr = pr + r_off
        if rr >= len(data): 
            continue
        for col in cols:
            if len(data[rr]) <= col:
                continue
            val = parse_float(data[rr][col])
            if val is None or val <= 0:
                continue
            is_y = flag_at(data, rr, col, y_base)
            is_s = flag_at(data, rr, col, s_base)
            if is_y and not is_s:
                return float(val), col, "YELLOW", rr
    return None, (cols[0] if cols else None), "NOT_FOUND_YELLOW", None

# === Стоимостные заголовки (умный выбор) ===
COST_POSITIVE_PAT = [
    r"\bcost\b", r"себест", r"себестоим", r"\bс/с\b", r"prime\s*cost"
]
COST_RUBLE_HINT_PAT = [r"руб", r"₽", r"\brub\b"]
COST_NEGATIVE_PAT = [
    r"\bold\b", r"стар", r"prev", r"предыд", r"архив", r"hist"
]
COST_BAD_CCY_PAT = [r"€", r"\beur\b", r"\$", r"\busd\b", r"cny", r"евро", r"доллар"]
COST_MINLIKE_PAT = [r"\bmin\b", r"\bмин\b", r"min15", r"1[.,]?5", r"порог", r"floor", r"target"]

def _score_cost_header(text: str) -> Optional[int]:
    if not isinstance(text, str): return None
    s = norm(text)
    score = 0
    if not any(re.search(p, s, flags=re.IGNORECASE) for p in COST_POSITIVE_PAT):
        return None  # вообще не «cost»-похоже — не рассматриваем
    score += 100
    if any(re.search(p, s, flags=re.IGNORECASE) for p in COST_RUBLE_HINT_PAT): score += 50
    if any(re.search(p, s, flags=re.IGNORECASE) for p in COST_NEGATIVE_PAT):  score -= 300
    if any(re.search(p, s, flags=re.IGNORECASE) for p in COST_BAD_CCY_PAT):   score -= 500  # режем €/$
    if any(re.search(p, s, flags=re.IGNORECASE) for p in COST_MINLIKE_PAT):   score -= 200
    # «идеальные» варианты — чисто Cost / Себестоимость
    if s in {"cost","себестоимость"}: score += 30
    return score

def find_cost_col_smart(data: List[List], pr: int, y_base: int) -> Tuple[int, Optional[str]]:
    """
    Находит ЛУЧШУЮ колонку Cost по хедеру: +RUB, −OLD/€/архив/мин15.
    Смотрим верхнюю строку и строку заголовков блока товара.
    Возвращает (col_idx0 | -1, header_text | None)
    """
    best_col, best_score, best_header = -1, -10**9, None
    for j in range(0, y_base):
        cand_parts = []
        top = data[0][j] if (len(data)>0 and len(data[0])>j) else None
        hdr = data[pr][j] if (pr < len(data) and len(data[pr])>j) else None
        if top: cand_parts.append(str(top))
        if hdr and hdr != top: cand_parts.append(str(hdr))
        if not cand_parts: continue
        header_text = " / ".join(cand_parts)
        sc = _score_cost_header(header_text)
        if sc is None: continue
        if sc > best_score:
            best_score, best_col, best_header = sc, j, header_text
    return best_col, best_header

def _norm_hdr(text: str) -> str:
    s = str(text or "")
    s = s.lower().replace("ё", "е")
    # убираем валюты/знаки и пунктуацию, схлопываем пробелы
    s = re.sub(r"[€$₽р.,;:()\\[\\]\\-]+", " ", s)
    s = s.replace("руб", " ").replace("eur", " ").replace("руд", " ")
    s = s.replace("/", "")   # с/стоим → сстоим
    s = re.sub(r"\s+", " ", s).strip()
    return s

def _hdr_has_token(text: str, tokens: set) -> bool:
    s = _norm_hdr(text)
    return any(tok in s for tok in tokens)

def _is_cost_header(text: str) -> bool:
    return isinstance(text, str) and (norm(text) in COST_HEADER_TOKENS)

def _is_vip_header(text: str) -> bool:
    return isinstance(text, str) and norm(text).replace(" ", "").startswith("vip")

def find_col_by_tokens(data: List[List], pr: int, y_base: int, tokens: set,
                       fallback_1based: Optional[int]=None) -> int:
    """
    1) ищем по верхней строке (глобальные заголовки),
    2) если не нашли — пробуем строку товара (локальные заголовки),
    3) если всё ещё нет — берём fallback.
    """
    def _scan_row(row) -> int:
        if not row: return -1
        col = -1
        for j in range(0, y_base):
            h = row[j] if j < len(row) else None
            if _hdr_has_token(h, tokens):
                col = j; break
        return col

    top = data[0] if data else []
    col = _scan_row(top)
    if col == -1 and pr < len(data):
        col = _scan_row(data[pr])

    if col == -1 and fallback_1based is not None:
        col = fallback_1based - 1
    return col

def find_cost_vip_cols(data: List[List], pr: int, y_base: int) -> Tuple[int, int]:
    cost_col = vip_col = -1
    headers = data[pr] if pr < len(data) else []
    for j in range(0, y_base):
        h = headers[j] if j < len(headers) else None
        if _is_cost_header(h):
            cost_col = j
        if _is_vip_header(h):
            vip_col = j
    if cost_col == -1 or vip_col == -1:
        top = data[0] if data else []
        for j in range(0, y_base):
            h = top[j] if j < len(top) else None
            if _is_cost_header(h) and cost_col == -1:
                cost_col = j
            if _is_vip_header(h) and vip_col == -1:
                vip_col = j
    if cost_col == -1:
        cost_col = FALLBACK_COST_COL_1B - 1
    if vip_col == -1:
        vip_col = FALLBACK_VIP_COL_1B - 1
    return cost_col, vip_col

# ========= Сбор цен рынка =========

def extract_market_prices_for_product(
    data: List[List], pr: int, y_base: int, require_both_flags: bool = True
) -> List[Tuple[str, float]]:
    out: List[Tuple[str, float]] = []
    rr = pr + 2
    while rr < len(data) and not is_intlike(data[rr][1] if len(data[rr]) > 1 else None):
        if rr + 1 >= len(data):
            break
        for j in range(0, y_base):
            nm = data[rr][j] if len(data[rr]) > j else None
            if nm is None:
                continue
            fh = flag_at(data, rr, j, y_base)       # жёлтый у ИМЕНИ
            fn = flag_at(data, rr + 1, j, y_base)   # жёлтый у ЦЕНЫ
            ok = (fh and fn) if require_both_flags else (fh or fn)
            if not ok:
                continue
            price = parse_price(data[rr + 1][j] if len(data[rr + 1]) > j else None)
            if price is None or price <= 0:
                continue
            nm_str = str(nm).strip()
            if is_market_supplier(nm_str):
                out.append((nm_str, float(price)))
        rr += 1
    return out

def extract_market_prices_for_product_f17(data: List[List], pr: int, y_base: int, s_base: int) -> List[Tuple[str, float]]:
    """
    Строгое правило:
      - берём ЦЕНУ только если ячейка цены жёлтая и НЕ зачёркнута
      - имя не должно быть зачёркнуто (жёлтизна имени — по желанию; оставим необязательной)
      - исключаем внутренние колонки: cost/cash/f17/vip/ндс/prime
    """
    out: List[Tuple[str, float]] = []
    rr = pr + 2
    INTERNAL_TOKENS = {"cost","кост","себестоимость","cash","кэш","нал","наличный","наличные","f17","ф17","vip","вип","ндс","prime"}

    def _is_internal(col_name: str) -> bool:
        ns = norm(col_name)
        return any(tok in ns for tok in INTERNAL_TOKENS)

    while rr < len(data) and not is_intlike(data[rr][1] if len(data[rr])>1 else None):
        if rr + 1 >= len(data):
            break
        for j in range(0, y_base):
            nm = data[rr][j] if len(data[rr]) > j else None
            if nm is None:
                continue

            nm_str = str(nm).strip()
            if _is_internal(nm_str):
                continue

            # активность ЯЧЕЙКИ ЦЕНЫ (жёлтая и НЕ зачёркнутая)
            y_price = flag_at(data, rr + 1, j, y_base)
            s_price = flag_at(data, rr + 1, j, s_base)
            if not (y_price and not s_price):
                continue

            # если имя зачёркнуто — пропускаем
            s_name = flag_at(data, rr, j, s_base)
            if s_name:
                continue

            price = parse_price(data[rr + 1][j] if len(data[rr + 1]) > j else None)
            if price is None or price <= 0:
                continue

            if is_market_supplier(nm_str):
                out.append((nm_str, float(price)))
        rr += 1
    return out

def extract_market_prices_for_product_zakup(data: List[List], pr: int, y_base: int) -> List[Tuple[str, float]]:
    """
    Zakup: берём цену, если жёлтым помечено имя ИЛИ цена.
    (это мягче, чем общий сбор, где могли требовать обе жёлтыми)
    Strike тут сознательно не проверяем.
    Исключаем поставщиков: белка, авангард, андрей (только для закупки).
    """
    out: List[Tuple[str, float]] = []
    rr = pr + 2
    while rr < len(data) and not is_intlike(data[rr][1] if len(data[rr]) > 1 else None):
        if rr + 1 >= len(data):
            break
        for j in range(0, y_base):
            nm = data[rr][j] if len(data[rr]) > j else None
            if nm is None:
                continue
            # жёлтое имя ИЛИ жёлтая цена
            fh = flag_at(data, rr, j, y_base)
            fn = flag_at(data, rr + 1, j, y_base)
            if not (fh or fn):
                continue
            price = parse_price(data[rr + 1][j] if len(data[rr + 1]) > j else None)
            if price is None or price <= 0:
                continue
            nm_str = str(nm).strip()
            if is_market_supplier(nm_str) and not is_zakup_excluded_supplier(nm_str):
                out.append((nm_str, float(price)))
        rr += 1
    return out

# ========= Вспомогалки для Cash =========

def insertion_rank(prices_sorted: List[float], x: float) -> int:
    arr = sorted(prices_sorted + [x])
    for i,v in enumerate(arr):
        if abs(v - x) < 1e-9: return i+1
    return len(arr)

# ======== CASH BLOCK — DO NOT MODIFY ========
# (ниже — функции и константы, отвечающие только за Cash)

def trim_by_gap(prices_sorted: List[float], GAP_THRESHOLD=_GAP_THRESHOLD_FROZEN) -> Tuple[List[float], Optional[str]]:
    n = len(prices_sorted)
    if n < 2: return prices_sorted, None
    if n == 2:
        lo, hi = prices_sorted[0], prices_sorted[1]
        gap = (hi - lo) / hi if hi>0 else 0.0
        if gap >= GAP_THRESHOLD: return [lo], "N2_GAP_GE40_DROP_MAX"
        return prices_sorted, None
    gaps = []
    for i in range(n-1):
        lo, hi = prices_sorted[i], prices_sorted[i+1]
        gap = (hi - lo)/hi if hi>0 else 0.0
        gaps.append((i, gap))
    i_max, max_gap = max(gaps, key=lambda t: t[1])
    if max_gap >= GAP_THRESHOLD:
        if i_max == n-2: return prices_sorted[:-1], "DROP_MAX_BY_GAP40"
        if i_max == 0:   return prices_sorted[1:],  "DROP_MIN_BY_GAP40"
    return prices_sorted, None

# ========= Cash анализ =========

def cash_choose_price_p2(
    competitor_prices: List[float],
    cost_val: Optional[float],
    cash_old: Optional[float],
    MIN_SUPPLIERS=_MIN_SUPPLIERS_FROZEN,
    CASH_MIN_MARGIN_PCT=_CASH_MIN_MARGIN_PCT_FROZEN,
) -> Tuple[Optional[float], str, Optional[int]]:
    """
    Возвращает: (Cash_new, Reason, AchievedRank)
    Логика P2-target + новое правило для двух цен на рынке:
      • После чистки цен если их ровно 2 → целимся в середину (mid = (p1+p2)/2).
      • Если Cost есть → Cash_new = max(mid, Cost×1.015).
      • Если Cost нет → Cash_new = mid; при наличии Cash_old — берём max(mid, Cash_old).
      • Если цен ≥3 → прежняя логика p2: Cash_new = max(p2, Cost×1.015) либо p2/old cash при отсутствии Cost.
    Всегда округляем вниз к 0.1.
    """
    steps = []
    if not competitor_prices:
        return None, "NO_MARKET", None

    base_sorted = sorted(competitor_prices)
    cleaned, gap_reason = trim_by_gap(base_sorted)
    if gap_reason:
        steps.append(gap_reason)
    if len(cleaned) < MIN_SUPPLIERS:
        steps.append(f"MARKET_SUPPLIERS_LT_{MIN_SUPPLIERS}")
        return None, " | ".join(steps), None

    # Ровно 2 цены → таргет в середину
    if len(cleaned) == 2:
        p1, p2 = cleaned[0], cleaned[1]
        mid = (p1 + p2) / 2.0
        steps.append("N2_TARGET_MID")
        if cost_val is None:
            if cash_old is not None:
                cand = max(mid, cash_old)
                steps.append("NO_COST_USE_MAX(CASH_OLD,MID)")
            else:
                cand = mid
                steps.append("NO_COST_USE_MID")
        else:
            min_15 = cost_val * (1.0 + CASH_MIN_MARGIN_PCT)
            cand = mid
            if cand < min_15:
                cand = min_15
                steps.append("MID_BELOW_COST_PLUS_1P5→RAISE_TO_MIN")
        cash_final = floor1(cand)
        steps.append("FLOOR1")
        achieved_rank = insertion_rank(base_sorted, cash_final)
        return cash_final, " | ".join(steps), achieved_rank

    # ≥3 цены → стандартный P2 таргет
    p2 = cleaned[1]
    steps.append("P2_TARGET")

    if cost_val is None:
        if cash_old is not None:
            cand = max(cash_old, p2)
            steps.append("NO_COST_USE_MAX(CASH_OLD,P2)")
        else:
            cand = p2
            steps.append("NO_COST_USE_P2")
    else:
        min_15 = cost_val * (1.0 + CASH_MIN_MARGIN_PCT)
        cand = p2
        if cash_old is not None and cash_old > p2 * 1.05:
            steps.append("CASH_OLD_GT_P2_PLUS_5PCT→SET_P2")
            cand = p2
        if cand < min_15:
            steps.append("P2_BELOW_COST_PLUS_1P5→RAISE_TO_MIN")
            cand = min_15

    cash_final = floor1(cand)
    steps.append("FLOOR1")
    achieved_rank = insertion_rank(base_sorted, cash_final)
    return cash_final, " | ".join(steps), achieved_rank

# ===== CASH-ONLY collectors =====
def cash_collect_market_prices(data: List[List], pr: int, y_base: int) -> List[Tuple[str, float]]:
    """
    CASH-ONLY: собираем пары (supplier, price) по правилу:
      - имя ИЛИ цена жёлтая (strike игнорируем),
      - поставщик проходит is_market_supplier,
      - parse_price берёт первый сегмент "123/..."
    НИ В КАКОМ другом модуле (F17/Zakup/VIP) эту функцию НЕ использовать.
    """
    out: List[Tuple[str, float]] = []
    rr = pr + 2
    while rr < len(data) and not is_intlike(data[rr][1] if len(data[rr]) > 1 else None):
        if rr + 1 >= len(data):
            break
        for j in range(0, y_base):
            nm = data[rr][j] if len(data[rr]) > j else None
            if nm is None:
                continue
            fh = flag_at(data, rr, j, y_base)       # жёлтое имя
            fn = flag_at(data, rr + 1, j, y_base)   # жёлтая цена
            if not (fh or fn):
                continue
            price = parse_price(data[rr + 1][j] if len(data[rr + 1]) > j else None)
            if price is None or price <= 0:
                continue
            nm_str = str(nm).strip()
            if is_market_supplier(nm_str):
                out.append((nm_str, float(price)))
        rr += 1
    return out

# ======== END CASH BLOCK ========

# ========= Zakup логика выбора цен =========
# Исключения для закупки (только для закупки, не для других цен)
ZAKUP_EXCLUDED_SUPPLIERS = {"белка", "belka", "авангард", "avangard", "андрей", "andrey", "andrei"}

def is_zakup_excluded_supplier(s: str) -> bool:
    """Проверяет, исключен ли поставщик для закупки"""
    if not isinstance(s, str) or not s.strip():
        return False
    ns = _normalize_name_for_match(s)
    return any(excluded in ns for excluded in ZAKUP_EXCLUDED_SUPPLIERS)

def pick_zakup_from_sorted_prices(sorted_prices: List[float]) -> Tuple[Optional[float], str]:
    """Выбор позиции из отсортированных цен для закупки"""
    n = len(sorted_prices)
    if n == 0:
        return None, "NO_MARKET"
    if n >= 5:
        return sorted_prices[2], "N5PLUS_TAKE_3RD"
    if n == 4:
        return sorted_prices[2], "N4_TAKE_3RD"
    if n == 3:
        return sorted_prices[1], "N3_TAKE_MEDIAN"
    if n == 2:
        return sorted_prices[1], "N2_TAKE_2ND"
    return sorted_prices[0], "N1_ONLY_PRICE_USED"

def analyze_cash_df(df: pd.DataFrame) -> pd.DataFrame:
    data = df.values.tolist()
    y_base = detect_flag_base(data)
    s_base = detect_strike_base(data)
    if y_base is None or s_base is None:
        raise ValueError("Не определены базовые сдвиги флагов (yellow/strike).")

    # первая товарная строка
    first_row = -1
    for i in range(min(50, len(data))):
        if is_intlike(data[i][1] if len(data[i])>1 else None):
            first_row = i; break
    if first_row == -1:
        raise ValueError("Не найден первый товар в первых 50 строках.")

    # найти столбец остатков по заголовку или по статистике шаблона
    stock_col = find_col_by_tokens(data, first_row, y_base, STOCK_HEADER_TOKENS, fallback_1based=None)
    if stock_col == -1:
        # эвристика: выбираем колонку с макс. числом «остаточных» строк в окне строк
        scan_top = min(len(data), first_row + 300)
        best_j, best_cnt = -1, -1
        for j in range(0, y_base):
            cnt = 0
            for r in range(first_row, scan_top):
                v = data[r][j] if len(data[r]) > j else None
                if isinstance(v, str) and STOCK_PATTERN.match(v):
                    cnt += 1
            if cnt > best_cnt:
                best_cnt, best_j = cnt, j
        stock_col = best_j

    # список товаров
    product_rows: List[int] = []
    r = first_row
    while r < len(data):
        if is_intlike(data[r][1] if len(data[r])>1 else None):
            product_rows.append(r); r += 1
            while r < len(data) and not is_intlike(data[r][1] if len(data[r])>1 else None):
                r += 1
        else:
            r += 1

    results=[]
    for pr in product_rows:
        code = data[pr][2] if len(data[pr])>2 else None
        name = extract_name_row(data, pr, 4)
        brand = guess_brand(name)

        # ----- Остатки (robust) -----
        stock_val = virtual_stock_val = None
        stock_raw = None

        total, virt, raw = read_stock_robust(data, pr, y_base, stock_col)
        if total is not None:
            stock_val = total
            virtual_stock_val = virt
            stock_raw = raw

        # === COST/Cash_old ===  (VIP-правило Cost)
        # Колонку Cost берём как в VIP
        cost_col, _vip_col_dummy = find_cost_vip_cols(data, pr, y_base)
        raw_cost = None
        cost_val = None
        cost_src = "NOT_FOUND_YELLOW"
        cost_row_idx = None

        if cost_col is not None and cost_col >= 0 and pr + 1 < len(data) and len(data[pr + 1]) > cost_col:
            tmp = parse_float(data[pr + 1][cost_col])
            raw_cost = float(tmp) if (tmp is not None and tmp > 0) else None
            is_y = flag_at(data, pr + 1, cost_col, y_base)  # жёлтая обязательна
            # В VIP strike НЕ проверяем
            if raw_cost is not None and is_y:
                cost_val = raw_cost
                cost_src = "YELLOW"
                cost_row_idx = pr + 1

        # Cash_old: CASH-ONLY — берём число из pr+1, если не зачёркнуто (жёлтизна не обязательна)
        cash_col = find_col_by_tokens(data, pr, y_base, CASH_HEADER_TOKENS, fallback_1based=None)
        cash_old = None
        if cash_col is not None and cash_col >= 0 and pr + 1 < len(data) and len(data[pr + 1]) > cash_col:
            v = parse_float(data[pr + 1][cash_col])
            is_s = flag_at(data, pr + 1, cash_col, s_base)  # не допускаем зачёркнутую
            if v is not None and v > 0 and not is_s:
                cash_old = float(v)

        # рынок
        market_pairs = cash_collect_market_prices(data, pr, y_base)
        market_prices = [p for (_,p) in market_pairs]
        market_sorted_pairs = sorted(market_pairs, key=lambda t: t[1])
        prices_sorted = sorted(market_prices)

        # Stock==0 → не ставим цену
        if stock_val == 0:
            results.append({
                "Code": code, "Name": name, "Brand": brand,
                "Stock": stock_val, "Virtual_Stock": virtual_stock_val,
                "Cost": cost_val, "Cash_old": cash_old,
                "Market_Count": len(prices_sorted),
                "Market_Prices": "; ".join(map(lambda x: str(round_half_down_0_01(x)), prices_sorted)) if prices_sorted else None,                                                            
                "Market_suppliers_used": "; ".join([f"{n}:{p}" for n,p in market_sorted_pairs]) if market_pairs else None,                                                                    
                "Cash_new": None, "Rank_num": None, "Target": None,
                "MinMargin": None, "RankFeasible": None,
                "Margin_to_Cost": None, "Is_Margin_OK": None, "Delta": None,
                "Cost_col_idx": cost_col,
                "Cost_src": cost_src,
                "Cost_row_idx": cost_row_idx,
                "Rule": "STOCK_EQ_0_SKIP", "Reason": "STOCK_EQ_0"
            }); continue


        # Нет рынка (или <2) → фоллбек
        if len(prices_sorted) < MIN_SUPPLIERS:
            min_cost15 = (cost_val*(1.0+CASH_MIN_MARGIN_PCT)) if (cost_val is not None) else None
            fallback=None; reason=f"NO_MARKET_OR_LT_{MIN_SUPPLIERS}_SUPPLIERS"
            if cash_old is not None:
                # удерживаем минимум к Cost+1.5% если есть Cost
                if min_cost15 is not None:
                    fallback = floor1(max(cash_old, floor1(min_cost15)))
                    reason += " | USE_CASH_OLD_AND_ENSURE_MIN_1P5"
                else:
                    fallback = floor1(cash_old)
                    reason += " | USE_CASH_OLD"
            elif cost_val is not None:
                fallback=floor1(cost_val*(1.0+CASH_MIN_MARGIN_PCT)); reason += " | USE_COST_PLUS_1P5"
            results.append({
                "Code": code, "Name": name, "Brand": brand,
                "Stock": stock_val, "Virtual_Stock": virtual_stock_val,
                "Cost": cost_val, "Cash_old": cash_old,
                "Market_Count": len(prices_sorted),
                "Market_Prices": "; ".join(map(lambda x: str(round_half_down_0_01(x)), prices_sorted)) if prices_sorted else None,                                                            
                "Market_suppliers_used": "; ".join([f"{n}:{p}" for n,p in market_sorted_pairs]) if market_pairs else None,                                                                    
                "Cash_new": fallback, "Rank_num": None, "Target": None,
                "MinMargin": floor1(min_cost15) if min_cost15 is not None else None, "RankFeasible": None,                                                                                    
                "Margin_to_Cost": (fallback - cost_val) if (fallback is not None and cost_val is not None) else None,                                                                         
                "Is_Margin_OK": (fallback is not None and cost_val is not None and (fallback - cost_val) >= cost_val*CASH_MIN_MARGIN_PCT) if fallback is not None and cost_val is not None else None,                                                                                          
                "Delta": (fallback - cash_old) if (fallback is not None and cash_old is not None) else None,                                                                                  
                "Cost_col_idx": cost_col,
                "Cost_src": cost_src,
                "Cost_row_idx": cost_row_idx,
                "Rule": "CASH_FALLBACK", "Reason": reason
            }); continue

        # P2-таргет
        cash_candidate, reason, rank = cash_choose_price_p2(prices_sorted, cost_val, cash_old)
        if cash_candidate is None:
            # подстраховка
            min_cost15 = (cost_val*(1.0+CASH_MIN_MARGIN_PCT)) if (cost_val is not None) else None
            fallback=None; reason_fb="P2_NONE_FALLBACK"
            if cash_old is not None:
                if min_cost15 is not None:
                    fallback=floor1(max(cash_old, floor1(min_cost15)))
                    reason_fb += " | USE_CASH_OLD_AND_ENSURE_MIN_1P5"
                else:
                    fallback=floor1(cash_old); reason_fb += " | USE_CASH_OLD"
            elif cost_val is not None:
                fallback=floor1(cost_val*(1.0+CASH_MIN_MARGIN_PCT)); reason_fb += " | USE_COST_PLUS_1P5"
            cash_candidate=fallback; reason=f"{reason} | {reason_fb}" if reason else reason_fb

        margin_abs = (cash_candidate - cost_val) if (cash_candidate is not None and cost_val is not None) else None
        is_margin_ok = (margin_abs is not None and cost_val is not None and margin_abs >= cost_val*CASH_MIN_MARGIN_PCT) if margin_abs is not None else None
        delta = (cash_candidate - cash_old) if (cash_candidate is not None and cash_old is not None) else None

        results.append({
            "Code": code, "Name": name, "Brand": brand,
            "Stock": stock_val, "Virtual_Stock": virtual_stock_val,
            "Cost": cost_val, "Cash_old": cash_old,
            "Market_Count": len(prices_sorted),
            "Market_Prices": "; ".join(map(lambda x: str(round_half_down_0_01(x)), prices_sorted)) if prices_sorted else None,                                                               
            "Market_suppliers_used": "; ".join([f"{n}:{p}" for n,p in market_sorted_pairs]) if market_pairs else None,                                                                       
            "Cash_new": cash_candidate, "Rank_num": rank,
            "Target": "P2",  # явная цель
            "MinMargin": floor1(cost_val*(1.0+CASH_MIN_MARGIN_PCT)) if cost_val is not None else None,                                                                                       
            "RankFeasible": None,
            "Margin_to_Cost": margin_abs, "Is_Margin_OK": is_margin_ok, "Delta": delta,
            "Cost_col_idx": cost_col,
            "Cost_src": cost_src,
            "Cost_row_idx": cost_row_idx,
            "Rule": "CASH_P2_TARGET", "Reason": reason
        })

    out = pd.DataFrame(results)

    # Приведение типов
    for c in ["Stock","Virtual_Stock","Cost","Cash_old","Cash_new","Rank_num","Margin_to_Cost",
              "Delta","Market_Count","MinMargin"]:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce")

    # «Рейтинг цены»: r из N
    def _fmt_rank(row):
        rank = row.get("Rank_num"); cnt = row.get("Market_Count")
        try: cnt_val = int(cnt) if pd.notna(cnt) else None
        except: cnt_val = None
        try: r_val = int(rank) if pd.notna(rank) else None
        except: r_val = None
        if cnt_val and cnt_val>0: return f"{r_val} из {cnt_val}" if r_val is not None else f"— из {cnt_val}"
        return None
    out["Рейтинг цены"] = out.apply(_fmt_rank, axis=1)

    # Min15 = floor1(Cost×1.015)
    out["Min15"] = out["Cost"].apply(lambda c: floor1(c*(1.0+CASH_MIN_MARGIN_PCT)) if pd.notna(c) else None)
    out["Min15"] = pd.to_numeric(out["Min15"], errors="coerce")

    # Удалить лишние колонки, если есть
    out.drop(columns=[c for c in ["Min15","MinMargin"] if c in out.columns], inplace=True, errors="ignore")

    # Определить имя колонки с новым Cash
    cash_new_candidates = ["Cash_new", "New_Cash", "CashNew", "Cash"]
    cash_new_col = next((c for c in cash_new_candidates if c in out.columns), None)

    # Дельта (в % к новому Cash): (Cash_new - Cost) / Cash_new * 100
    if cash_new_col and "Cost" in out.columns:
        _cash_new_num = pd.to_numeric(out[cash_new_col], errors="coerce")
        _cost_num = pd.to_numeric(out["Cost"], errors="coerce")
        _pct = (_cash_new_num - _cost_num) / _cash_new_num * 100.0
        out["Delta_Cost_to_CashNew_%"] = _pct.apply(lambda v: round_half_down_0_01(v) if pd.notna(v) else None)
        out["Delta_Cost_to_CashNew_%"] = pd.to_numeric(out["Delta_Cost_to_CashNew_%"], errors="coerce")

    # Удаляем Brand из вывода, если присутствует
    if "Brand" in out.columns:
        out.drop(columns=["Brand"], inplace=True)

    preferred = [
        "Code","Name",
        "Stock","Virtual_Stock",
        "Cost","Cash_old","Cash_new","Delta","Delta_Cost_to_CashNew_%",
        "Рейтинг цены","Margin_to_Cost","Market_suppliers_used",
        "Target","Rule","Reason"
    ]
    preferred_present = [c for c in preferred if c in out.columns]
    tail = [c for c in out.columns if c not in preferred_present + ["Market_Count","Rank_num"]]
    out = out[preferred_present + tail]
    return out

# ========= F17 анализ =========

def choose_f17_price_p1(competitor_prices: List[float], cost_val: Optional[float], f17_old: Optional[float]) -> Tuple[Optional[float], str, Optional[int]]:
    """
    Динамический undercut (0.49 если p1..9, иначе 0.50), soft-cap на повышения,
    floor к 0.50, .99 только из целого, no-upmove против F17_old.
    """
    steps = []
    if not competitor_prices:
        return None, "NO_MARKET", None

    base_sorted = sorted(competitor_prices)
    cleaned, gap_reason = trim_by_gap(base_sorted)
    if gap_reason:
        steps.append(gap_reason)
    if len(cleaned) == 0:
        steps.append("MARKET_EMPTY_AFTER_TRIM")
        return None, " | ".join(steps), None

    if cost_val is None:
        steps.append("NO_COST_SKIP")
        return None, " | ".join(steps), None

    p1 = cleaned[0]
    steps.append("P1_TARGET")
    step = calc_undercut_step(p1)
    steps.append(f"UNDERCUT_STEP={step:.2f}")

    cand = target_by_market_softcap(p1, cost_val, f17_old)
    if cand is None:
        steps.append("P1_INFEASIBLE_MIN_ABS")
        return None, " | ".join(steps), None

    rank = insertion_rank(base_sorted, cand)
    return cand, " | ".join(steps), rank

def analyze_f17_df(df: pd.DataFrame) -> pd.DataFrame:
    data = df.values.tolist()
    y_base = detect_flag_base(data)
    s_base = detect_strike_base(data)
    if y_base is None or s_base is None:
        raise ValueError("Не определены базовые сдвиги флагов (yellow/strike).")

    # первая товарная строка
    first_row = -1
    for i in range(min(50, len(data))):
        if is_intlike(data[i][1] if len(data[i])>1 else None):
            first_row = i; break
    if first_row == -1:
        raise ValueError("Не найден первый товар в первых 50 строках.")

    # найти столбец остатков по заголовку или по статистике шаблона
    stock_col = find_col_by_tokens(data, first_row, y_base, STOCK_HEADER_TOKENS, fallback_1based=None)
    if stock_col == -1:
        # эвристика: выбираем колонку с макс. числом «остаточных» строк в окне строк
        scan_top = min(len(data), first_row + 300)
        best_j, best_cnt = -1, -1
        for j in range(0, y_base):
            cnt = 0
            for r in range(first_row, scan_top):
                v = data[r][j] if len(data[r]) > j else None
                if isinstance(v, str) and STOCK_PATTERN.match(v):
                    cnt += 1
            if cnt > best_cnt:
                best_cnt, best_j = cnt, j
        stock_col = best_j

    # список товаров
    product_rows: List[int] = []
    r = first_row
    while r < len(data):
        if is_intlike(data[r][1] if len(data[r])>1 else None):
            product_rows.append(r); r += 1
            while r < len(data) and not is_intlike(data[r][1] if len(data[r])>1 else None):
                r += 1
        else:
            r += 1

    results=[]
    for pr in product_rows:
        code = data[pr][2] if len(data[pr])>2 else None
        name = extract_name_row(data, pr, 4)

        # ----- Остаток (robust) -----
        stock_val = virtual_stock_val = None
        stock_raw = None

        total, virt, raw = read_stock_robust(data, pr, y_base, stock_col)
        if total is not None:
            stock_val = total
            virtual_stock_val = virt
            stock_raw = raw

        # COST/Cash/F17_old
        # VIP-логика выбора колонки Cost
        cost_col, _vip_col_dummy = find_cost_vip_cols(data, pr, y_base)
        cash_col = find_col_by_tokens(data, pr, y_base, CASH_HEADER_TOKENS, fallback_1based=None)
        f17_col = find_col_by_tokens(data, pr, y_base, F17_HEADER_TOKENS, fallback_1based=None)

        cost_val = None
        if pr + 1 < len(data) and cost_col is not None and cost_col >= 0 and len(data[pr + 1]) > cost_col:
            c = parse_float(data[pr + 1][cost_col])
            if c is not None and c > 0:
                is_y = flag_at(data, pr + 1, cost_col, y_base)  # жёлтая обязательна
                if is_y:
                    cost_val = float(c)

        cash_old=None
        if cash_col is not None and cash_col>=0 and pr+1 < len(data) and len(data[pr+1])>cash_col:
            v = parse_float(data[pr+1][cash_col])
            is_y = flag_at(data, pr+1, cash_col, y_base)
            is_s = flag_at(data, pr+1, cash_col, s_base)
            if v is not None and v>0 and is_y and not is_s:
                cash_old = float(v)
            else:
                cash_old = None

        # F17_old: брать ТОЛЬКО если клетка жёлтая и НЕ зачёркнута
        f17_old = None
        if pr + 1 < len(data) and f17_col is not None and f17_col >= 0 and len(data[pr + 1]) > f17_col:
            v = parse_float(data[pr + 1][f17_col])
            is_y = flag_at(data, pr + 1, f17_col, y_base)   # жёлтая
            is_s = flag_at(data, pr + 1, f17_col, s_base)   # зачёркнута
            if v is not None and v > 0 and is_y and not is_s:
                f17_old = float(v)
            else:
                f17_old = None

        # рынок (СТРОГИЙ — как в f17.py)
        market_pairs = extract_market_prices_for_product_f17(data, pr, y_base, s_base)
        market_prices = [p for (_, p) in market_pairs]
        market_sorted_pairs = sorted(market_pairs, key=lambda t: t[1])
        prices_sorted = sorted(market_prices)

        # Stock==0 → не ставим цену
        if stock_val == 0:
            f17_old_raw = parse_float(data[pr + 1][f17_col]) if (f17_col is not None and pr + 1 < len(data) and len(data[pr + 1]) > f17_col) else None
            f17_old_flag = ("OK" if (f17_old is not None) else "SKIPPED")
            results.append({
                "Code": code, "Name": name,
                "Stock": stock_val, "Virtual_Stock": virtual_stock_val,
                "Cost": cost_val, "Cash": cash_old, "F17_old": f17_old,
                "F17_old_raw": f17_old_raw, "F17_old_flag": f17_old_flag,
                "Market_Count": len(prices_sorted),
                "Market_Prices": "; ".join(map(lambda x: str(round_half_down_0_01(x)), prices_sorted)) if prices_sorted else None,
                "Market_suppliers_used": "; ".join([f"{n}:{p}" for n, p in market_sorted_pairs]) if market_pairs else None,
                "F17_new": None, "Rank_num": None,
                "Rule": "STOCK_EQ_0_SKIP", "Reason": "STOCK_EQ_0"
            })
            continue

        # Нет рынка → держим F17_old (если валиден)
        if len(prices_sorted) == 0:
            f17_old_raw = parse_float(data[pr + 1][f17_col]) if (f17_col is not None and pr + 1 < len(data) and len(data[pr + 1]) > f17_col) else None
            f17_old_flag = ("OK" if (f17_old is not None) else "SKIPPED")
            if f17_old is not None:
                results.append({
                    "Code": code, "Name": name,
                    "Stock": stock_val, "Virtual_Stock": virtual_stock_val,
                    "Cost": cost_val, "Cash": cash_old, "F17_old": f17_old,
                    "F17_old_raw": f17_old_raw, "F17_old_flag": f17_old_flag,
                    "Market_Count": 0, "Market_Prices": None, "Market_suppliers_used": None,
                    "F17_new": f17_old, "Rank_num": None,
                    "Rule": "NO_MARKET_USE_F17_OLD",
                    "Reason": "NO_MARKET | USE_F17_OLD (valid yellow & not strike)"
                })
            else:
                results.append({
                    "Code": code, "Name": name,
                    "Stock": stock_val, "Virtual_Stock": virtual_stock_val,
                    "Cost": cost_val, "Cash": cash_old, "F17_old": f17_old,
                    "F17_old_raw": f17_old_raw, "F17_old_flag": f17_old_flag,
                    "Market_Count": 0, "Market_Prices": None, "Market_suppliers_used": None,
                    "F17_new": None, "Rank_num": None,
                    "Rule": "NO_MARKET", "Reason": "NO_MARKET | NO_VALID_F17_OLD"
                })
            continue

        # Чистка анти-выбросов и p1
        cleaned, gap_reason = trim_by_gap(prices_sorted)
        p1_local = cleaned[0] if len(cleaned) >= 1 else None
        step_eff = calc_undercut_step(p1_local) if p1_local is not None else 0.50

        # --- TIGHTEN_GAP_UPMOVE: подтяжка вверх, если старый #1 и gap ≥ 1.00 ---
        f17_candidate = None
        rule_val = None
        reason = None
        ALLOW_UPMOVE = False

        if (GAP_TIGHTEN_ENABLED and f17_old is not None and cost_val is not None and p1_local is not None):
            gap = p1_local - f17_old  # насколько мы ниже рынка
            if gap >= GAP_TIGHTEN_TRIGGER:
                target_max = p1_local - step_eff
                cand = target_max
                if ROUND_TO_HALF:
                    cand = round_to_0_50(cand)
                if ROUND_ENDING_99:
                    cand = apply_ending_99_if_integer(cand)

                # допускаем небольшое отклонение от минимума при подтяжке
                TIGHTEN_MIN_DELTA = -0.05
                if cand >= cost_val + TIGHTEN_MIN_DELTA and cand > f17_old + 1e-9:
                    # CAP по Cash: не дороже Cash, НО не ниже старой F17 в режиме подтяжки
                    if cash_old is not None:
                        min_allowed = cost_val + 0.50
                        if cand > cash_old and cash_old >= min_allowed - 1e-9:
                            cand_cap = cash_old
                            # ключевая защита: не опускаем ниже старой F17
                            if f17_old is not None and cand_cap < f17_old:
                                cand = f17_old
                            else:
                                cand = cand_cap

                            # косметика .99 (как у тебя)
                            if ROUND_ENDING_99:
                                cand = apply_ending_99_if_integer(cand)

                            # держим ≤ p1 - step (страховка)
                            if p1_local is not None and cand > (p1_local - step_eff) + 1e-9:
                                cand = p1_local - step_eff
                                if ROUND_TO_HALF:
                                    cand = round_to_0_50(cand)
                                if ROUND_ENDING_99:
                                    cand = apply_ending_99_if_integer(cand)

                        # и всегда соблюдаем минимум к себестоимости
                        if cand < min_allowed - 1e-9:
                            cand = min_allowed

                    f17_candidate = float(Decimal(str(cand)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
                    rule_val = "TIGHTEN_GAP_UPMOVE"
                    reason = f"OLD_IS_#1 | GAP={gap:.2f}≥{GAP_TIGHTEN_TRIGGER} | SET_TO_P1_MINUS_{step_eff:.2f}"
                    ALLOW_UPMOVE = True  # разрешаем повышение в этом кейсе

        # --- если подтяжка не сработала, основной расчёт ---
        if f17_candidate is None:
            # если старая уже #1 и не ниже мин.порога — держим старую
            keep_old = False
            if f17_old is not None and p1_local is not None:
                if (p1_local - f17_old) >= 0.01 and (cost_val is None or (f17_old - cost_val) >= GRANDFATHER_MIN_MARGIN - 1e-9):
                    keep_old = True

            if keep_old:
                f17_candidate = f17_old
                rule_val = "KEEP_OLD_STABLE"
                reason = "OLD_IS_STILL_#1 | DO_NOT_RAISE"
                rank = insertion_rank(sorted(prices_sorted), f17_candidate)
            else:
                # Новый кандидат под p1
                f17_candidate, reason, rank = choose_f17_price_p1(prices_sorted, cost_val, f17_old)
                rule_val = "F17_FROM_P1"

                # CAP по Cash (не дороже Cash), при этом не нарушаем минимум Cost+0.50 и #1
                if f17_candidate is not None and cash_old is not None:
                    min_allowed = cost_val + 0.50 if cost_val is not None else None
                    if (min_allowed is None) or (cash_old >= min_allowed - 1e-9):
                        if f17_candidate > cash_old:
                            f17_candidate = cash_old
                            if ROUND_ENDING_99:
                                f17_candidate = apply_ending_99_if_integer(f17_candidate)
                            # сохраняем #1
                            if p1_local is not None and f17_candidate > (p1_local - step_eff) + 1e-9:
                                f17_candidate = p1_local - step_eff
                                if ROUND_TO_HALF:
                                    f17_candidate = round_to_0_50(f17_candidate)
                                if ROUND_ENDING_99:
                                    f17_candidate = apply_ending_99_if_integer(f17_candidate)

                # Если решения нет — проверяем infeasible по минимуму
                if f17_candidate is None and p1_local is not None and cost_val is not None:
                    min_allowed = cost_val + 0.50
                    infeasible_by_market = (min_allowed > (p1_local - step_eff) + 1e-9)
                    if infeasible_by_market and SKIP_OLD_WHEN_MARKET_INFEASIBLE:
                        # рынок слишком низкий → ничего не ставим (и НЕ откатываемся)
                        rule_val = "SKIP_INFEASIBLE_MARKET_TOO_LOW"
                        reason = (reason + " | P1_INFEASIBLE_MIN_ABS") if reason else "P1_INFEASIBLE_MIN_ABS"
                    elif f17_old is not None and p1_local is not None:
                        # если не infeasible — оставим старую при условии #1 и порогов
                        if (f17_old <= (p1_local - step_eff) + 1e-9 and
                            (f17_old >= cost_val + 0.50 - 1e-9) and
                            (f17_old <= cost_val + 3.00 + 1e-9)):
                            f17_candidate = f17_old
                            rule_val = "KEEP_OLD_STABLE"
                            reason = (reason + " | OLD_STILL_#1") if reason else "OLD_STILL_#1"

        # Глобальный запрет повышений (кроме явной подтяжки)
        if (f17_candidate is not None and f17_old is not None and
            f17_candidate > f17_old + 1e-9 and not ALLOW_UPMOVE):
            f17_candidate = f17_old
            rule_val = "NO_UPMOVE_KEEP_OLD"
            reason = (reason + " | NO_UPMOVE") if reason else "NO_UPMOVE"

        margin_abs = (f17_candidate - cost_val) if (f17_candidate is not None and cost_val is not None) else None
        margin_pct = round((margin_abs / cost_val * 100.0), 2) if (margin_abs is not None and cost_val not in (None, 0)) else None
        delta = (f17_candidate - f17_old) if (f17_candidate is not None and f17_old is not None) else None

        f17_old_raw = parse_float(data[pr + 1][f17_col]) if (f17_col is not None and pr + 1 < len(data) and len(data[pr + 1]) > f17_col) else None
        f17_old_flag = ("OK" if (f17_old is not None) else "SKIPPED")

        results.append({
            "Code": code, "Name": name,
            "Stock": stock_val, "Virtual_Stock": virtual_stock_val,
            "Cost": cost_val, "Cash": cash_old, "F17_old": f17_old,
            "F17_old_raw": f17_old_raw, "F17_old_flag": f17_old_flag,
            "Market_Count": len(prices_sorted),
            "Market_Prices": "; ".join(map(lambda x: str(round_half_down_0_01(x)), prices_sorted)) if prices_sorted else None,
            "Market_suppliers_used": "; ".join([f"{n}:{p}" for n, p in market_sorted_pairs]) if market_pairs else None,
            "F17_new": f17_candidate, "Rank_num": rank if 'rank' in locals() else None,
            "Margin_to_Cost": margin_abs, "Margin_to_Cost_%": margin_pct, "Delta_F17": delta,
            "Rule": rule_val, "Reason": reason
        })

    out = pd.DataFrame(results)

    # Приведение типов
    for c in ["Stock","Virtual_Stock","Cost","Cash","F17_old","F17_new","Rank_num","Margin_to_Cost","Margin_to_Cost_%","Delta_F17","Market_Count"]:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce")
    
    # Форматирование Delta_F17 как в cash анализе (с round_half_down_0_01)
    if "Delta_F17" in out.columns:
        out["Delta_F17"] = out["Delta_F17"].apply(lambda x: round_half_down_0_01(x) if pd.notna(x) else None)

    # «Рейтинг цены»: r из N
    def _fmt_rank(row):
        rank = row.get("Rank_num"); cnt = row.get("Market_Count")
        try: cnt_val = int(cnt) if pd.notna(cnt) else None
        except: cnt_val = None
        try: r_val = int(rank) if pd.notna(rank) else None
        except: r_val = None
        if cnt_val and cnt_val>0: return f"{r_val} из {cnt_val}" if r_val is not None else f"— из {cnt_val}"
        return None
    out["Рейтинг цены"] = out.apply(_fmt_rank, axis=1)

    preferred = [
        "Code","Name",
        "Stock","Virtual_Stock",
        "Cost","Cash","F17_old","F17_new","Delta_F17","Margin_to_Cost_%",
        "Рейтинг цены","Market_suppliers_used",
        "Rule","Reason"
    ]
    preferred_present = [c for c in preferred if c in out.columns]
    tail = [c for c in out.columns if c not in preferred_present + ["Market_Count","Rank_num","Margin_to_Cost"]]
    out = out[preferred_present + tail]
    return out

# ========= Zakup Analysis =========

def pick_zakup_from_sorted_prices(sorted_prices: List[float]) -> Tuple[Optional[float], str]:
    n = len(sorted_prices)
    if n == 0:
        return None, "NO_MARKET"
    if n >= 5:
        return sorted_prices[2], "N5PLUS_TAKE_3RD"
    if n == 4:
        return sorted_prices[2], "N4_TAKE_3RD"
    if n == 3:
        return sorted_prices[1], "N3_TAKE_MEDIAN"
    if n == 2:
        return sorted_prices[1], "N2_TAKE_2ND"
    return sorted_prices[0], "N1_ONLY_PRICE_USED"


def analyze_zakup_df(df: pd.DataFrame) -> pd.DataFrame:
    data = df.values.tolist()
    y_base = detect_flag_base(data)
    s_base = detect_strike_base(data)
    if y_base is None or s_base is None:
        raise ValueError("Не определены базовые сдвиги флагов (yellow/strike).")

    # первая товарная строка
    first_row = -1
    for i in range(min(50, len(data))):
        if is_intlike(data[i][1] if len(data[i]) > 1 else None):
            first_row = i
            break
    if first_row == -1:
        raise ValueError("Не найден первый товар в первых 50 строках.")

    # найти столбец остатков по заголовку или по статистике шаблона
    stock_col = find_col_by_tokens(data, first_row, y_base, STOCK_HEADER_TOKENS, fallback_1based=None)
    if stock_col == -1:
        # эвристика: колонка с максимумом строк в формате "остатков"
        scan_top = min(len(data), first_row + 300)
        best_j, best_cnt = -1, -1
        for j in range(0, y_base):
            cnt = 0
            for r in range(first_row, scan_top):
                v = data[r][j] if len(data[r]) > j else None
                if isinstance(v, str) and STOCK_PATTERN.match(v):
                    cnt += 1
            if cnt > best_cnt:
                best_cnt, best_j = cnt, j
        stock_col = best_j

    # список товаров
    product_rows: List[int] = []
    r = first_row
    while r < len(data):
        if is_intlike(data[r][1] if len(data[r]) > 1 else None):
            product_rows.append(r)
            r += 1
            while r < len(data) and not is_intlike(data[r][1] if len(data[r]) > 1 else None):
                r += 1
        else:
            r += 1

    results = []
    for pr in product_rows:
        code = data[pr][2] if len(data[pr]) > 2 else None
        name = extract_name_row(data, pr, 4)

        # ----- Остаток (robust) -----
        stock_val = virtual_stock_val = None
        stock_raw = None

        total, virt, raw = read_stock_robust(data, pr, y_base, stock_col)
        if total is not None:
            stock_val = total
            virtual_stock_val = virt
            stock_raw = raw

        # колонки cost / zakup_old
        # VIP-логика выбора колонки Cost
        cost_col, _vip_col_dummy = find_cost_vip_cols(data, pr, y_base)
        zakup_col = find_col_by_tokens(data, pr, y_base, ZAKUP_HEADER_TOKENS, fallback_1based=None)

        # COST (VIP-правило): жёлтая ячейка, strike не проверяем
        cost_val = None
        if pr + 1 < len(data) and cost_col is not None and cost_col >= 0 and len(data[pr + 1]) > cost_col:
            c = parse_float(data[pr + 1][cost_col])
            if c is not None and c > 0:
                is_y = flag_at(data, pr + 1, cost_col, y_base)
                if is_y:
                    cost_val = float(c)

        # Zakup_old (берём как число, без флаговой жёсткости)
        zakup_old = None
        if zakup_col is not None and zakup_col >= 0 and pr + 1 < len(data) and len(data[pr + 1]) > zakup_col:
            z = parse_float(data[pr + 1][zakup_col])
            zakup_old = float(z) if (z is not None and z > 0) else None

        # рынок (Zakup: мягкий сбор «имя или цена жёлтая»)
        market_pairs = extract_market_prices_for_product_zakup(data, pr, y_base)
        market_prices_sorted = sorted([p for (_, p) in market_pairs])
        market_sorted_pairs = sorted(market_pairs, key=lambda t: t[1])

        # --- спец-правила по стоку/вирту ---
        # Stock=0 & VS=35 → если есть Zakup_old — держим его, иначе пропуск
        if (stock_val is not None and stock_val == 0) and (virtual_stock_val == 35):
            if zakup_old is not None:
                zakup_new = floor1(zakup_old)
                delta = round(zakup_new - zakup_old, 2) if zakup_new is not None else None
                change = "Y" if (delta is not None and abs(delta) >= 0.01) else "N"
                results.append({
                    "Code": code, "Name": name, "Stock": stock_val, "Virtual_Stock": virtual_stock_val,
                    "Cost": cost_val, "Zakup_old": zakup_old,
                    "Market_Count": len(market_prices_sorted),
                    "Market_Prices": "; ".join(map(lambda x: str(round_half_down_0_01(x)), market_prices_sorted)) if market_prices_sorted else None,
                    "Market_suppliers_used": "; ".join([f"{n}:{p}" for n, p in market_sorted_pairs]) if market_pairs else None,
                    "Market_Selected": None, "Chosen_Index": None,
                    "Zakup_new": zakup_new, "Delta": delta, "Change": change,
                    "Rule": "STOCK0_VS35_USE_OLD", "Reason": "STOCK0_VS35_HAS_OLD", "Stock_raw": stock_raw,
                })
            else:
                results.append({
                    "Code": code, "Name": name, "Stock": stock_val, "Virtual_Stock": virtual_stock_val,
                    "Cost": cost_val, "Zakup_old": zakup_old,
                    "Market_Count": len(market_prices_sorted),
                    "Market_Prices": "; ".join(map(lambda x: str(round_half_down_0_01(x)), market_prices_sorted)) if market_prices_sorted else None,
                    "Market_suppliers_used": "; ".join([f"{n}:{p}" for n, p in market_sorted_pairs]) if market_pairs else None,
                    "Market_Selected": None, "Chosen_Index": None,
                    "Zakup_new": None, "Delta": None, "Change": "N",
                    "Rule": "STOCK0_VS35_SKIP_NO_OLD", "Reason": "NO_MARKS_STOCK0_VS35_NO_OLD", "Stock_raw": stock_raw,
                })
            continue

        # Stock=0 & VS=0 → пропуск
        if (stock_val is not None and stock_val == 0) and (virtual_stock_val == 0):
            results.append({
                "Code": code, "Name": name, "Stock": stock_val, "Virtual_Stock": virtual_stock_val,
                "Cost": cost_val, "Zakup_old": zakup_old,
                "Market_Count": len(market_prices_sorted),
                "Market_Prices": "; ".join(map(lambda x: str(round_half_down_0_01(x)), market_prices_sorted)) if market_prices_sorted else None,
                "Market_suppliers_used": "; ".join([f"{n}:{p}" for n, p in market_sorted_pairs]) if market_pairs else None,
                "Market_Selected": None, "Chosen_Index": None,
                "Zakup_new": None, "Delta": None, "Change": "N",
                "Rule": "STOCK0_VS0_SKIP", "Reason": "NO_MARKS_STOCK0_VS0", "Stock_raw": stock_raw,
            })
            continue

        # Stock>0 и рынка нет → Zakup_old
        if (stock_val is not None and stock_val > 0) and len(market_prices_sorted) == 0:
            zakup_new = floor1(zakup_old) if zakup_old is not None else None
            delta = round(zakup_new - zakup_old, 2) if (zakup_new is not None and zakup_old is not None) else None
            change = "Y" if ((delta is not None and abs(delta) >= 0.01) or (zakup_new is not None and zakup_old is None)) else "N"
            results.append({
                "Code": code, "Name": name, "Stock": stock_val, "Virtual_Stock": virtual_stock_val,
                "Cost": cost_val, "Zakup_old": zakup_old,
                "Market_Count": 0, "Market_Prices": None, "Market_suppliers_used": None,
                "Market_Selected": None, "Chosen_Index": None,
                "Zakup_new": zakup_new, "Delta": delta, "Change": change,
                "Rule": "STOCK_GT0_NO_MARKET_USE_OLD", "Reason": "STOCK_GT0_PRICE_MUST_EXIST",
                "Stock_raw": stock_raw,
            })
            continue

        # --- анти-выбросы и выбор по рынку ---
        price_chosen = None
        reason = None

        # n=2: разрыв ≥40% → берём минимальную
        if len(market_prices_sorted) == 2:
            lo, hi = market_prices_sorted[0], market_prices_sorted[1]
            gap = (hi - lo) / hi if hi > 0 else 0.0
            if gap >= GAP_THRESHOLD:
                price_chosen = lo
                reason = "N2_GAP_GE40_TAKE_MIN"

        # n≥3: выброс по соседям
        if price_chosen is None and len(market_prices_sorted) >= 3:
            gaps = []
            for i in range(len(market_prices_sorted) - 1):
                lo = market_prices_sorted[i]
                hi = market_prices_sorted[i + 1]
                gap = (hi - lo) / hi if hi > 0 else 0.0
                gaps.append((i, gap))
            if gaps:
                i_max, max_gap = max(gaps, key=lambda t: t[1])
                if max_gap >= GAP_THRESHOLD:
                    if i_max == len(market_prices_sorted) - 2:
                        trimmed = market_prices_sorted[:-1]
                        price_chosen, _ = pick_zakup_from_sorted_prices(trimmed)
                        reason = "ADJ_GAP_GE40_DROP_MAX_THEN_RULE"
                    elif i_max == 0:
                        trimmed = market_prices_sorted[1:]
                        price_chosen, _ = pick_zakup_from_sorted_prices(trimmed)
                        reason = "ADJ_GAP_GE40_DROP_MIN_THEN_RULE"

        # обычный выбор по рынку, если ещё не выбрали
        if price_chosen is None:
            price_chosen, base_reason = pick_zakup_from_sorted_prices(market_prices_sorted)
            if reason is None:
                reason = base_reason

        # выбранный поставщик/индекс в отсортированной паре
        chosen_index = None
        market_selected = None
        if price_chosen is not None:
            for idx, (n, p) in enumerate(market_sorted_pairs):
                if abs(p - price_chosen) < 1e-9:
                    chosen_index = idx + 1
                    market_selected = f"{n}:{p}"
                    break

        # ==== защита (после выбора рынка) ====
        # Пол: при наличии стока и валидном Cost — не ниже Cost×1.015
        if (stock_val is not None and stock_val > 0) and (cost_val is not None) and (price_chosen is not None):
            if price_chosen < cost_val:
                price_chosen = cost_val * 1.015
                reason = (reason + " | " if reason else "") + "RAISED_TO_COST_PLUS_1_5_STOCK_GT0"

        # Потолок: не выше Old×1.15 (если old есть)
        if (price_chosen is not None) and (zakup_old is not None):
            cap = zakup_old * (1.0 + MARKET_ABOVE_OLD_CAP_PCT)
            if price_chosen > cap:
                price_chosen = cap
                reason = (reason + " | " if reason else "") + "CAPPED_OLD_PLUS_15"

        zakup_new = floor1(price_chosen) if price_chosen is not None else None
        rule = "STANDARD"

        delta = None
        change = "N"
        if (zakup_new is not None) and (zakup_old is not None):
            delta = round(zakup_new - zakup_old, 2)
            change = "Y" if abs(delta) >= 0.01 else "N"
        elif zakup_new is not None and zakup_old is None:
            change = "Y"

        results.append({
            "Code": code, "Name": name, "Stock": stock_val, "Virtual_Stock": virtual_stock_val,
            "Cost": cost_val, "Zakup_old": zakup_old,
            "Market_Count": len(market_prices_sorted),
            "Market_Prices": "; ".join(map(lambda x: str(round_half_down_0_01(x)), market_prices_sorted)) if market_prices_sorted else None,
            "Market_suppliers_used": "; ".join([f"{n}:{p}" for n, p in market_sorted_pairs]) if market_pairs else None,
            "Market_Selected": market_selected, "Chosen_Index": chosen_index,
            "Zakup_new": zakup_new, "Delta": delta, "Change": change,
            "Rule": rule, "Reason": reason, "Stock_raw": stock_raw,
        })

    out = pd.DataFrame(results)
    for c in ["Stock", "Virtual_Stock", "Cost", "Zakup_old", "Zakup_new", "Delta", "Chosen_Index"]:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce")

    # Маржа к себестоимости в процентах: (Zakup_new - Cost) / Cost * 100
    def _margin_pct(z, cost):
        if pd.isna(z) or pd.isna(cost) or cost == 0:
            return None
        return round_half_down_0_01((z - cost) / cost * 100.0)
    if "Zakup_new" in out.columns and "Cost" in out.columns:
        out["Margin_to_Cost_%"] = [_margin_pct(z, c) for z, c in zip(out["Zakup_new"], out["Cost"])]
        out["Margin_to_Cost_%"] = pd.to_numeric(out["Margin_to_Cost_%"], errors="coerce")

    # Порядок столбцов (под f17-стиль) + Market_Prices в конце по желанию
    preferred = [
        "Code", "Name",
        "Stock", "Virtual_Stock",
        "Cost", "Zakup_old", "Zakup_new", "Delta", "Margin_to_Cost_%",
        "Market_Selected", "Chosen_Index",
        "Market_suppliers_used",
        "Rule", "Reason"
    ]
    preferred_present = [c for c in preferred if c in out.columns]
    tail = [c for c in out.columns if c not in preferred_present + ["Market_Prices"]]
    out = out[preferred_present + tail + (["Market_Prices"] if "Market_Prices" in out.columns else [])]
    return out

# ======== VIP BLOCK — DO NOT MODIFY ========
# (ниже — функции и константы, отвечающие только за VIP)

# ===== VIP: референс/рынок (особые правила: Парель/Парк Марк/Шамбор) =====
# --- VIP-only ---
PAREL_MARK_PATTERNS_VIP = [
    r"\bпарел[ьи]\s*ма", r"\bпарел[ьи]\s*марк",
    r"\bparel+\s*ma",    r"\bparel+\s*mark",
    r"\bparell+\s*ma",   r"\bparell+\s*mark",
]

# Базовый список рынка (мягкие матчинги)
MARKET_SUPPLIERS = [
    "Шанталь С...",
    "борики братья - гульден 14",
    "борики братья — гульден 14",
    "борики братья",
    "гульден 14",
    "владимир",
    "владимир спец",
    "дима американец",
    "лужники ирина (пав.1.6)",
    "лужники ирина",
    "пав.1.6",
    "настя марина",
    "натс",
    "парк",  # обычный Парк (но НЕ «парк марк»)
    "саша черный",
    "тихом",
    "тоня",
    "сергей",
    "альянс групп",
]

# Явные исключения рынка
EXCLUDED_MARKET_SUPPLIERS = ["владимир марк", "vladimir mark"]

# Регэксп-исключения
EXCLUDED_MARKET_PATTERNS = [
    r"^\s*владимир\s*\.\.\.\s*$",
    r"(?:^|[\s(\[{])владимир\s*\.\.\.(?:$|[\s)\]}:,-])",
]

def _vip_norm(s: str) -> str:
    ns = norm(s).replace("—", "-").replace("…", "...")
    ns = re.sub(r"\.{2,}", "...", ns)
    ns = re.sub(r"[^\w\s().-:]", " ", ns)
    ns = re.sub(r"\s+", " ", ns).strip()
    return ns

def _vip_word_boundary_contains(haystack: str, needle: str) -> bool:
    pat = r"(?:^|[\s().-:])" + re.escape(needle) + r"(?:$|[\s().-:])"
    return re.search(pat, haystack, flags=re.IGNORECASE) is not None

def _is_parel_mark_vip(ns: str) -> bool:
    return any(re.search(p, ns, flags=re.IGNORECASE) for p in PAREL_MARK_PATTERNS_VIP)

def _is_bare_parel_vip(ns: str) -> bool:
    if re.search(r"\bпарел[ьи]\b(?!\s*ма)", ns, flags=re.IGNORECASE): return True
    if re.search(r"\bparel{1,2}\b(?!\s*ma)", ns, flags=re.IGNORECASE): return True
    return False

def is_ref_supplier_vip(s: str) -> bool:
    if not isinstance(s, str): return False
    ns = _vip_norm(s)
    if _is_bare_parel_vip(ns): return False
    if "парк марк" in ns or "park mark" in ns: return True
    if "шамбор" in ns and ("ндс" in ns or "без нд" in ns): return True
    if _is_parel_mark_vip(ns): return True
    return False

def is_market_supplier_vip(s: str) -> bool:
    """
    Рынок для VIP: исключаем «Парк Марк» и любую «Парель», остальное — мягкие правила.
    """
    if not isinstance(s, str) or not s.strip(): return False
    ns = _vip_norm(s)

    # исключаем референс «Парк Марк»
    if "парк марк" in ns or "park mark" in ns: return False
    # любую Парель исключаем из рынка
    if _is_bare_parel_vip(ns) or _is_parel_mark_vip(ns): return False

    # явные исключения
    for bad in EXCLUDED_MARKET_SUPPLIERS:
        if _vip_word_boundary_contains(ns, _vip_norm(bad)):
            return False
    for pat in EXCLUDED_MARKET_PATTERNS:
        if re.search(pat, ns, flags=re.IGNORECASE):
            return False

    # мягкие совпадения рынка
    if re.search(r"\bлужники\s*и(?:[^a-zа-я]|$)", ns, flags=re.IGNORECASE): return True
    if re.search(r"\bнастя\s*мар(?:ина|\b|[^a-zа-я])", ns, flags=re.IGNORECASE): return True
    if re.search(r"(?:^|[\s().:;\-])нат[сc](?:$|[\s().:;\-])", ns, flags=re.IGNORECASE): return True
    if re.search(r"(?:^|[\s().:;\-])парк(?:$|[\s().:;\-])", ns, flags=re.IGNORECASE): return True
    if re.search(r"\bальянс\s*гр", ns, flags=re.IGNORECASE): return True

    # «Владимир» ровно словом — рынок
    if re.fullmatch(r"\s*владимир\s*", ns, flags=re.IGNORECASE): return True

    # базовый список
    for key in MARKET_SUPPLIERS:
        if _vip_word_boundary_contains(ns, _vip_norm(key)):
            return True

    return False

# ========= VIP Analysis =========

def _vip_is_cost_header(text: str) -> bool:
    if not isinstance(text, str):
        return False
    t = norm(text).replace(" ", "")
    return t.startswith("cost") or t.startswith("кост") or t.startswith("себест")

def _vip_is_vip_header(text: str) -> bool:
    if not isinstance(text, str):
        return False
    t = norm(text).replace(" ", "")
    return t.startswith("vip")

def find_cost_vip_stock_cell(data: List[List], pr: int, y_base: int) -> Tuple[int, int, Optional[Tuple[int,int]]]:
    cost_col = vip_col = -1
    headers = data[pr] if pr < len(data) else []

    for j in range(0, y_base):
        h = headers[j] if j < len(headers) else None
        if _vip_is_cost_header(h):
            cost_col = j
        if _vip_is_vip_header(h):
            vip_col = j

    # stock: ищем в строке товара или 1-2 строки ниже
    stock_cell = None
    for r_off in (0, 1, 2):
        rr = pr + r_off
        if rr >= len(data):
            break
        for j in range(0, y_base):
            v = data[rr][j] if len(data[rr]) > j else None
            if isinstance(v, str) and STOCK_PATTERN.match(v):
                stock_cell = (rr, j)
                break
        if stock_cell:
            break

    return cost_col, vip_col, stock_cell


def analyze_vip_df(df: pd.DataFrame, VIP_FALLBACK_COL=_VIP_FALLBACK_COL_FROZEN) -> pd.DataFrame:
    data = df.values.tolist()
    y_base = detect_flag_base(data)
    s_base = detect_strike_base(data)
    if y_base is None or s_base is None:
        raise ValueError("Не удалось определить базовые сдвиги флагов (yellow/strike).")

    # первая товарная строка
    first_row = -1
    for i in range(min(50, len(data))):
        if is_intlike(data[i][1] if len(data[i]) > 1 else None):
            first_row = i
            break
    if first_row == -1:
        raise ValueError("Не найден первый товар в первых 50 строках.")

    # найти столбец остатков (по заголовку или по статистике)
    stock_col = find_col_by_tokens(data, first_row, y_base, {"остаток","остатки","stock"}, fallback_1based=None)
    if stock_col == -1:
        scan_top = min(len(data), first_row + 300)
        best_j, best_cnt = -1, -1
        for j in range(0, y_base):
            cnt = 0
            for r in range(first_row, scan_top):
                v = data[r][j] if len(data[r]) > j else None
                if isinstance(v, str) and STOCK_PATTERN.match(v):
                    cnt += 1
            if cnt > best_cnt:
                best_cnt, best_j = cnt, j
        stock_col = best_j

    # список товаров
    product_rows: List[int] = []
    r = first_row
    while r < len(data):
        if is_intlike(data[r][1] if len(data[r]) > 1 else None):
            product_rows.append(r)
            r += 1
            while r < len(data) and not is_intlike(data[r][1] if len(data[r]) > 1 else None):
                r += 1
        else:
            r += 1

    results = []
    for pr in product_rows:
        code = data[pr][2] if len(data[pr]) > 2 else None
        name = extract_name_row(data, pr, 4)
        cost_col, vip_col = find_cost_vip_cols(data, pr, y_base)

        # ----- Остаток (robust) -----
        stock_val = virtual_stock_val = None
        stock_raw = None

        total, virt, raw = read_stock_robust(data, pr, y_base, stock_col)
        if total is not None:
            stock_val = total
            virtual_stock_val = virt
            stock_raw = raw

        # COST: только если ячейка жёлтая; позже можем снять при Stock=0
        raw_cost = parse_float(data[pr + 1][cost_col]) if pr + 1 < len(data) and cost_col != -1 else None
        raw_cost = None if (raw_cost is None or raw_cost <= 0) else float(raw_cost)
        cost_yellow = flag_at(data, pr + 1, cost_col, y_base) if cost_col != -1 else False
        cost_exists = (raw_cost is not None) and cost_yellow
        cost_used = raw_cost if cost_exists else None
        forced_no_cost = False
        cost_original = raw_cost
        cost_changed = False

        # Глобальные стопы
        if virtual_stock_val == 35:
            results.append({
                "Code": code, "B": "", "Name": name,
                "Stock": stock_val, "Virtual_Stock": virtual_stock_val,
                "Cost": None, "Cost_Yellow": "N",
                "Old_VIP": None, "New_VIP": None, "Delta": None, "Change": "N",
                "Rule": "VIRTUAL_STOCK_35_SKIP",
                "X_ref": None, "X_ref_supplier": None, "Reason": "VIRTUAL_STOCK_35_SKIP",
                "Market_suppliers_used": None, "Market_suppliers_all": None,
                "Market_suppliers_all_count": 0, "Stock_raw": stock_raw,
                "Cost_Forced_NoCost": "N", "Cost_Original": cost_original, "Cost_Changed": "N",
            })
            continue

        if (stock_val is not None and stock_val <= 0) and (virtual_stock_val is not None and virtual_stock_val == 0):
            results.append({
                "Code": code, "B": "", "Name": name,
                "Stock": stock_val, "Virtual_Stock": virtual_stock_val,
                "Cost": None, "Cost_Yellow": "N",
                "Old_VIP": None, "New_VIP": None, "Delta": None, "Change": "N",
                "Rule": "STOCK_0_AND_VIRT_0_SKIP",
                "X_ref": None, "X_ref_supplier": None, "Reason": "STOCK_0_AND_VIRTUAL_0_SKIP",
                "Market_suppliers_all": None, "Market_suppliers_all_count": 0, "Stock_raw": stock_raw,
                "Cost_Forced_NoCost": "N", "Cost_Original": cost_original, "Cost_Changed": "N",
            })
            continue

        colB_note = "старая цена (нет цены)" if not cost_exists else ""

        # Если реального стока нет — снимаем Cost
        no_real_stock = (stock_val is not None and stock_val <= 0)
        if no_real_stock and cost_exists:
            forced_no_cost = True
            cost_changed = True
            cost_exists = False
            cost_used = None
            cost_yellow = False
            colB_note = (colB_note + " | COST снят из-за Stock=0").strip(" |")

        # OLD VIP: только если жёлтый и не зачёркнут
        oldVIP_raw = parse_float(data[pr + 1][vip_col]) if pr + 1 < len(data) and vip_col != -1 else None
        vip_yellow = flag_at(data, pr + 1, vip_col, y_base) if vip_col != -1 else False
        vip_strike = flag_at(data, pr + 1, vip_col, s_base) if vip_col != -1 else False
        oldVIP_exists = (oldVIP_raw is not None) and vip_yellow and (not vip_strike)
        oldVIP_used = float(oldVIP_raw) if oldVIP_exists else None

        # Рынок: имя ИЛИ цена жёлтые
        ref_prices = []
        market_prices = []
        rr = pr + 2
        while rr < len(data) and not is_intlike(data[rr][1] if len(data[rr]) > 1 else None):
            if rr + 1 >= len(data):
                break
            for j in range(0, y_base):
                nm = data[rr][j] if len(data[rr]) > j else None
                if nm is None:
                    continue
                fh = flag_at(data, rr, j, y_base)
                fn = flag_at(data, rr + 1, j, y_base)
                if not (fh or fn):
                    continue
                price = parse_price(data[rr + 1][j] if len(data[rr + 1]) > j else None)
                if price is None or price <= 0:
                    continue
                nm_str = str(nm).strip()
                if is_ref_supplier_vip(nm_str):
                    ref_prices.append((nm_str, price))
                if is_market_supplier_vip(nm_str):
                    market_prices.append((nm_str, price))
            rr += 1

        market_all_sorted = sorted(market_prices, key=lambda t: t[1])
        market_all_str = "; ".join([f"{n}:{p}" for n, p in market_all_sorted]) if market_all_sorted else None
        market_all_cnt = len(market_all_sorted)

        # === ЛОГИКА ===
        if cost_exists:
            if not ref_prices:
                ref = None
                xref_supplier = None
                if oldVIP_used is not None:
                    newVIP = oldVIP_used
                    reason = "NO_REF_PRICE_KEEP_OLD"
                else:
                    guard_candidate = ceil1(cost_used * 1.01) if cost_used is not None else None
                    newVIP = floor1(guard_candidate) if guard_candidate is not None else None
                    reason = "NO_REF_PRICE_GUARD_1pct" if newVIP is not None else "NO_REF_PRICE_NO_COST"
            else:
                pmin = min([p for _, p in ref_prices])
                xref_suppliers = [n for (n, p) in ref_prices if p == pmin]
                xref_supplier = "; ".join(sorted(set(map(str, xref_suppliers)))) if xref_suppliers else None

                ref_raw = pmin / 1.2
                ref = round_half_down_0_01(ref_raw)
                guard = ceil1(cost_used * 1.01) if cost_used is not None else None
                cands = []
                if ref is not None:
                    cands.append(float(ref))
                if guard is not None:
                    cands.append(float(guard))
                newVIP = max(cands) if cands else oldVIP_used
                if newVIP is not None and cands:
                    newVIP = floor1(float(newVIP))
                reason = None

            delta = None if (newVIP is None or oldVIP_used is None) else round(newVIP - oldVIP_used, 2)
            change = "Y" if (delta is not None and abs(delta) >= 0.01) else "N"
            results.append({
                "Code": code, "B": colB_note, "Name": name,
                "Stock": stock_val, "Virtual_Stock": virtual_stock_val,
                "Cost": cost_used, "Cost_Yellow": "Y" if cost_yellow else "N",
                "Old_VIP": oldVIP_used, "New_VIP": newVIP, "Delta": delta, "Change": change,
                "Rule": "STANDARD_COST_PRESENT",
                "X_ref": ref if 'ref' in locals() else None,
                "X_ref_supplier": xref_supplier if 'xref_supplier' in locals() else None,
                "Reason": reason,
                "Market_suppliers_used": None, "Market_suppliers_all": market_all_str,
                "Market_suppliers_all_count": market_all_cnt,
                "Stock_raw": stock_raw, "Cost_Forced_NoCost": "Y" if forced_no_cost else "N",
                "Cost_Original": cost_original, "Cost_Changed": "Y" if cost_changed else "N",
            })
            continue

        # ---- НЕТ COST ----

        # 1) Есть реальный сток и Old_VIP — держим старый
        if (stock_val is not None and stock_val > 0) and (oldVIP_used is not None):
            results.append({
                "Code": code, "B": colB_note, "Name": name,
                "Stock": stock_val, "Virtual_Stock": virtual_stock_val,
                "Cost": None, "Cost_Yellow": "N",
                "Old_VIP": oldVIP_used, "New_VIP": oldVIP_used, "Delta": 0.0, "Change": "N",
                "Rule": "NO_COST_STOCK_POS_KEEP_OLD",
                "X_ref": None, "X_ref_supplier": None, "Reason": "NO_COST_STOCK_GT0_KEEP_OLDVIP",
                "Market_suppliers_used": None, "Market_suppliers_all": market_all_str,
                "Market_suppliers_all_count": market_all_cnt,
                "Stock_raw": stock_raw,
                "Cost_Forced_NoCost": "Y" if forced_no_cost else "N",
                "Cost_Original": cost_original, "Cost_Changed": "Y" if cost_changed else "N",
            })
            continue

        # 2) Считаем X_ref и рынок
        pmin_ref = min([p for _, p in ref_prices]) if ref_prices else None
        if ref_prices and pmin_ref is not None:
            xref_suppliers = [n for (n, p) in ref_prices if p == pmin_ref]
            xref_supplier = "; ".join(sorted(set(map(str, xref_suppliers)))) if xref_suppliers else None
        else:
            xref_supplier = None

        ref_raw  = (pmin_ref / 1.2) if pmin_ref is not None else None
        ref      = round_half_down_0_01(ref_raw) if ref_raw is not None else None

        market_list = market_all_sorted
        newVIP = None
        reason = None
        market_used = None

        if ref is not None:
            ref_cmp = ref
            market_lower = [(n, p) for (n, p) in market_list if p < ref_cmp]

            if len(market_lower) >= 1:
                newVIP = floor1(float(ref))
                reason = "PASS_PMIN_USE_XREF"
                market_used = "; ".join([f"{n}:{p}" for n, p in sorted(market_lower, key=lambda t: t[1])])
            else:
                if (stock_val is not None and stock_val > 0) and (oldVIP_used is None) and len(market_list) >= 1:
                    n_min, p_min = market_list[0]
                    newVIP = floor1(float(p_min) * 1.01)
                    reason = "NO_PMIN_PASS_USE_MIN_MARKET_PLUS1pct"
                    market_used = f"{n_min}:{p_min}"
                else:
                    if len(market_list) >= 2:
                        n2, p2 = market_list[1]
                        newVIP = floor1(float(p2) * 1.015)
                        reason = "NO_MARKET_BELOW_USE_SECOND_PLUS1_5"
                        market_used = f"{n2}:{p2}"
                    else:
                        reason = "NO_MARKET_PRICE_BELOW_CANDIDATE"
        else:
            if len(market_list) >= 2:
                n2, p2 = market_list[1]
                newVIP = floor1(float(p2) * 1.20)
                reason = "NO_REF_USE_SECOND_PLUS20"
                market_used = f"{n2}:{p2}"
            elif len(market_list) == 1:
                n1, p1 = market_list[0]
                newVIP = None
                reason = "NO_REF_ONLY_ONE_MARKET_SKIP"
                market_used = f"{n1}:{p1}"
            else:
                newVIP = None
                reason = "NO_REF_PRICE_OR_NO_MARKET"
                market_used = None

        delta = None if (newVIP is None or oldVIP_used is None) else round(newVIP - oldVIP_used, 2)
        results.append({
            "Code": code, "B": colB_note, "Name": name,
            "Stock": stock_val, "Virtual_Stock": virtual_stock_val,
            "Cost": None, "Cost_Yellow": "N",
            "Old_VIP": oldVIP_used, "New_VIP": newVIP, "Delta": delta,
            "Change": "Y" if (delta is not None and abs(delta) >= 0.01) else "N" if newVIP is not None else "N",
            "Rule": "NO_COST_MARKET_REF_RULE",
            "X_ref": ref, "X_ref_supplier": xref_supplier, "Reason": reason,
            "Market_suppliers_used": market_used, "Market_suppliers_all": market_all_str,
            "Market_suppliers_all_count": market_all_cnt,
            "Stock_raw": stock_raw, "Cost_Forced_NoCost": "Y" if forced_no_cost else "N",
            "Cost_Original": cost_original, "Cost_Changed": "Y" if cost_changed else "N",
        })

    out = pd.DataFrame(results)

    # Приведение типов
    for c in ["Cost","Old_VIP","New_VIP","Delta","X_ref","Stock","Virtual_Stock","Cost_Original"]:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce")
    if "Market_suppliers_all_count" in out.columns:
        out["Market_suppliers_all_count"] = pd.to_numeric(out["Market_suppliers_all_count"], errors="coerce").fillna(0).astype(int)

    # Вставить X_ref_supplier сразу после X_ref
    cols = list(out.columns)
    if "X_ref" in cols and "X_ref_supplier" in cols:
        cols.remove("X_ref_supplier")
        try:
            xi = cols.index("X_ref") + 1
            cols = cols[:xi] + ["X_ref_supplier"] + cols[xi:]
            out = out[cols]
        except ValueError:
            pass

    # Маржа к себестоимости, %
    def _margin_pct(new_vip, cost):
        if pd.isna(new_vip) or pd.isna(cost) or cost == 0:
            return None
        return round_half_down_0_01((new_vip - cost) / cost * 100.0)
    if "New_VIP" in out.columns and "Cost" in out.columns:
        out["Margin_to_Cost_%"] = [ _margin_pct(n, c) for n, c in zip(out["New_VIP"], out["Cost"]) ]
        out["Margin_to_Cost_%"] = pd.to_numeric(out["Margin_to_Cost_%"], errors="coerce")

    # Порядок колонок
    preferred = [
        "Code", "Name",
        "Stock", "Virtual_Stock",
        "Cost", "Old_VIP", "New_VIP", "Delta", "Margin_to_Cost_%",
        "X_ref", "X_ref_supplier",
        "Market_suppliers_all_count", "Market_suppliers_all", "Market_suppliers_used",
        "Rule", "Reason", "B"
    ]
    preferred_present = [c for c in preferred if c in out.columns]
    tail = [c for c in out.columns if c not in preferred_present]
    out = out[preferred_present + tail]
    return out

# ========= Excel Export Functions =========

def export_cash_to_excel(cash_df: pd.DataFrame) -> BytesIO:
    """Экспорт Cash анализа в Excel с условным форматированием"""
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        cash_df.to_excel(writer, sheet_name='Cash_Analysis', index=False)
        
        # Получаем workbook и worksheet
        workbook = writer.book
        worksheet = writer.sheets['Cash_Analysis']
        
        # Замораживаем первую строку
        worksheet.freeze_panes = 'A2'
        
        # Условное форматирование
        try:
            cols_map = {name: i + 1 for i, name in enumerate(cash_df.columns)}
            last_row = cash_df.shape[0] + 1
            
            # Градиент для Delta_Cost_to_CashNew_%
            if "Delta_Cost_to_CashNew_%" in cols_map:
                col_letter = get_column_letter(cols_map["Delta_Cost_to_CashNew_%"])
                rng = f"{col_letter}2:{col_letter}{last_row}"
                worksheet.conditional_formatting.add(
                    rng,
                    ColorScaleRule(
                        start_type='min', start_color='FFF8696B',
                        mid_type='percentile', mid_value=50, mid_color='FFFFEB84',
                        end_type='max', end_color='FF63BE7B'
                    )
                )
            
            # Градиент для Delta
            if "Delta" in cols_map:
                dcol = get_column_letter(cols_map["Delta"])
                drng = f"{dcol}2:{dcol}{last_row}"
                worksheet.conditional_formatting.add(
                    drng,
                    ColorScaleRule(
                        start_type='num', start_value=-1, start_color='FFF4CCCC',
                        mid_type='num', mid_value=0, mid_color='FFD9D9D9',
                        end_type='num', end_value=1, end_color='FF93C47D'
                    )
                )
            
            # Cash_new: синий если изменилась цена
            if "Cash_new" in cols_map and "Cash_old" in cols_map:
                cash_new_col = get_column_letter(cols_map["Cash_new"])
                cash_old_col = get_column_letter(cols_map["Cash_old"])
                cash_new_range = f"{cash_new_col}2:{cash_new_col}{last_row}"
                
                worksheet.conditional_formatting.add(
                    cash_new_range,
                    FormulaRule(
                        formula=[f"AND({cash_new_col}2<>{cash_old_col}2,{cash_new_col}2<>\"\")"],
                        fill=PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                    )
                )
            
            # Пустые значения: светло-серый
            empty_cols = ["Cost", "Cash_old", "Cash_new"]
            for col_name in empty_cols:
                if col_name in cols_map:
                    col_idx = cols_map[col_name]
                    col_range = f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{last_row}"
                    
                    worksheet.conditional_formatting.add(
                        col_range,
                        CellIsRule(
                            operator="equal", formula=["\"\""],
                            fill=PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                        )
                    )
        except Exception as e:
            st.warning(f"Ошибка условного форматирования: {e}")
    
    output.seek(0)
    return output

def export_f17_to_excel(f17_df: pd.DataFrame) -> BytesIO:
    """Экспорт F17 анализа в Excel с условным форматированием"""
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        f17_df.to_excel(writer, sheet_name='F17_Analysis', index=False)
        
        # Получаем workbook и worksheet
        workbook = writer.book
        worksheet = writer.sheets['F17_Analysis']
        
        # Замораживаем первую строку
        worksheet.freeze_panes = 'A2'
        
        # Условное форматирование
        try:
            cols_map = {name: i + 1 for i, name in enumerate(f17_df.columns)}
            last_row = f17_df.shape[0] + 1
            
            # Градиент для Delta_F17
            if "Delta_F17" in cols_map:
                delta_col = get_column_letter(cols_map["Delta_F17"])
                delta_range = f"{delta_col}2:{delta_col}{last_row}"
                
                worksheet.conditional_formatting.add(
                    delta_range,
                    ColorScaleRule(
                        start_type="num", start_value=-1, start_color="FF6B6B",
                        mid_type="num", mid_value=0, mid_color="D3D3D3",
                        end_type="num", end_value=1, end_color="90EE90"
                    )
                )
            
            # F17_new: синий если изменилась цена
            if "F17_new" in cols_map and "F17_old" in cols_map:
                f17_new_col = get_column_letter(cols_map["F17_new"])
                f17_old_col = get_column_letter(cols_map["F17_old"])
                f17_new_range = f"{f17_new_col}2:{f17_new_col}{last_row}"
                
                worksheet.conditional_formatting.add(
                    f17_new_range,
                    FormulaRule(
                        formula=[f"AND({f17_new_col}2<>{f17_old_col}2,{f17_new_col}2<>\"\")"],
                        fill=PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                    )
                )
            
            # Margin_to_Cost_% — 3-цветная шкала (красный→жёлтый→зелёный)
            if "Margin_to_Cost_%" in cols_map:
                margin_col = get_column_letter(cols_map["Margin_to_Cost_%"])
                margin_range = f"{margin_col}2:{margin_col}{last_row}"
                
                worksheet.conditional_formatting.add(
                    margin_range,
                    ColorScaleRule(
                        start_type='min', start_color='FFF8696B',
                        mid_type='percentile', mid_value=50, mid_color='FFFFEB84',
                        end_type='max', end_color='FF63BE7B'
                    )
                )
            
            # Пустые значения: светло-серый
            empty_cols = ["Cost", "Cash", "F17_old", "F17_new"]
            for col_name in empty_cols:
                if col_name in cols_map:
                    col_idx = cols_map[col_name]
                    col_range = f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{last_row}"
                    
                    worksheet.conditional_formatting.add(
                        col_range,
                        CellIsRule(
                            operator="equal", formula=["\"\""],
                            fill=PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                        )
                    )
        except Exception as e:
            st.warning(f"Ошибка условного форматирования: {e}")
    
    output.seek(0)
    return output

def export_zakup_to_excel(zakup_df: pd.DataFrame) -> BytesIO:
    """Экспорт Zakup анализа в Excel с условным форматированием"""
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        zakup_df.to_excel(writer, sheet_name='Zakup_Analysis', index=False)
        
        # Получаем workbook и worksheet
        workbook = writer.book
        worksheet = writer.sheets['Zakup_Analysis']
        
        # Замораживаем первую строку
        worksheet.freeze_panes = 'A2'
        
        # Условное форматирование
        try:
            cols_map = {name: i + 1 for i, name in enumerate(zakup_df.columns)}
            last_row = zakup_df.shape[0] + 1
            
            # Градиент для Delta
            if "Delta" in cols_map:
                delta_col = get_column_letter(cols_map["Delta"])
                delta_range = f"{delta_col}2:{delta_col}{last_row}"
                
                worksheet.conditional_formatting.add(
                    delta_range,
                    ColorScaleRule(
                        start_type="num", start_value=-1, start_color="FF6B6B",
                        mid_type="num", mid_value=0, mid_color="D3D3D3",
                        end_type="num", end_value=1, end_color="90EE90"
                    )
                )
            
            # Zakup_new: синий если изменилась цена
            if "Zakup_new" in cols_map and "Zakup_old" in cols_map:
                zakup_new_col = get_column_letter(cols_map["Zakup_new"])
                zakup_old_col = get_column_letter(cols_map["Zakup_old"])
                zakup_new_range = f"{zakup_new_col}2:{zakup_new_col}{last_row}"
                
                worksheet.conditional_formatting.add(
                    zakup_new_range,
                    FormulaRule(
                        formula=[f"AND({zakup_new_col}2<>{zakup_old_col}2,{zakup_new_col}2<>\"\")"],
                        fill=PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                    )
                )
            
            # Margin_to_Cost_% — 3-цветная шкала (красный→жёлтый→зелёный)
            if "Margin_to_Cost_%" in cols_map:
                margin_col = get_column_letter(cols_map["Margin_to_Cost_%"])
                margin_range = f"{margin_col}2:{margin_col}{last_row}"
                
                worksheet.conditional_formatting.add(
                    margin_range,
                    ColorScaleRule(
                        start_type='min', start_color='FFF8696B',
                        mid_type='percentile', mid_value=50, mid_color='FFFFEB84',
                        end_type='max', end_color='FF63BE7B'
                    )
                )
            
            # Пустые значения: светло-серый
            empty_cols = ["Cost", "Zakup_old", "Zakup_new"]
            for col_name in empty_cols:
                if col_name in cols_map:
                    col_idx = cols_map[col_name]
                    col_range = f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{last_row}"
                    
                    worksheet.conditional_formatting.add(
                        col_range,
                        CellIsRule(
                            operator="equal", formula=["\"\""],
                            fill=PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                        )
                    )
        except Exception as e:
            st.warning(f"Ошибка условного форматирования: {e}")
    
    output.seek(0)
    return output

# ======== END VIP BLOCK ========

def export_vip_to_excel(vip_df: pd.DataFrame) -> BytesIO:
    """Экспорт VIP анализа в Excel с условным форматированием"""
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        vip_df.to_excel(writer, sheet_name='VIP_Analysis', index=False)
        
        # Получаем workbook и worksheet
        workbook = writer.book
        worksheet = writer.sheets['VIP_Analysis']
        
        # Замораживаем первую строку
        worksheet.freeze_panes = 'A2'
        
        # Условное форматирование
        try:
            cols_map = {name: i + 1 for i, name in enumerate(vip_df.columns)}
            last_row = vip_df.shape[0] + 1
            
            # Градиент для Delta
            if "Delta" in cols_map:
                delta_col = get_column_letter(cols_map["Delta"])
                delta_range = f"{delta_col}2:{delta_col}{last_row}"
                
                worksheet.conditional_formatting.add(
                    delta_range,
                    ColorScaleRule(
                        start_type="num", start_value=-1, start_color="FF6B6B",
                        mid_type="num", mid_value=0, mid_color="D3D3D3",
                        end_type="num", end_value=1, end_color="90EE90"
                    )
                )
            
            # New_VIP: синий если изменилась цена
            if "New_VIP" in cols_map and "Old_VIP" in cols_map:
                new_vip_col = get_column_letter(cols_map["New_VIP"])
                old_vip_col = get_column_letter(cols_map["Old_VIP"])
                new_vip_range = f"{new_vip_col}2:{new_vip_col}{last_row}"
                
                worksheet.conditional_formatting.add(
                    new_vip_range,
                    FormulaRule(
                        formula=[f"AND({new_vip_col}2<>{old_vip_col}2,{new_vip_col}2<>\"\")"],
                        fill=PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                    )
                )
            
            # Margin_to_Cost_% — 3-цветная шкала (красный→жёлтый→зелёный)
            if "Margin_to_Cost_%" in cols_map:
                margin_col = get_column_letter(cols_map["Margin_to_Cost_%"])
                margin_range = f"{margin_col}2:{margin_col}{last_row}"
                
                worksheet.conditional_formatting.add(
                    margin_range,
                    ColorScaleRule(
                        start_type='min', start_color='FFF8696B',
                        mid_type='percentile', mid_value=50, mid_color='FFFFEB84',
                        end_type='max', end_color='FF63BE7B'
                    )
                )
            
            # Пустые значения: светло-серый
            empty_cols = ["Cost", "Old_VIP", "New_VIP"]
            fill_gray = PatternFill(fill_type="solid", fgColor="FFD9D9D9")
            for col_name in empty_cols:
                if col_name in cols_map:
                    col_letter = get_column_letter(cols_map[col_name])
                    col_range = f"{col_letter}2:{col_letter}{last_row}"
                    
                    worksheet.conditional_formatting.add(
                        col_range,
                        FormulaRule(
                            formula=[f'OR(ISBLANK({col_letter}2),LEN(TRIM({col_letter}2))=0,{col_letter}2="")'],
                            fill=fill_gray,
                            stopIfTrue=False
                        )
                    )
            
            # Выделить всю строку красным, если есть остаток (Stock>0), но New_VIP пустой
            if "Stock" in cols_map and "New_VIP" in cols_map:
                stock_letter = get_column_letter(cols_map["Stock"])
                newvip_letter = get_column_letter(cols_map["New_VIP"])
                last_col_letter = get_column_letter(len(vip_df.columns))
                
                # Формула для выделения строки
                formula = f'AND(${stock_letter}2>0, LEN(TRIM(${newvip_letter}2&""))=0)'
                row_range = f"A2:{last_col_letter}{last_row}"
                
                worksheet.conditional_formatting.add(
                    row_range,
                    FormulaRule(
                        formula=[formula],
                        fill=PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                    )
                )
        except Exception as e:
            st.warning(f"Ошибка условного форматирования: {e}")
    
    output.seek(0)
    return output

def export_all_analyses_to_excel(cash_df: pd.DataFrame, f17_df: pd.DataFrame, 
                                zakup_df: pd.DataFrame, vip_df: pd.DataFrame) -> BytesIO:
    """Экспорт всех анализов в один Excel файл с 4 листами"""
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Записываем все анализы на отдельные листы
        cash_df.to_excel(writer, sheet_name='Cash_Analysis', index=False)
        f17_df.to_excel(writer, sheet_name='F17_Analysis', index=False)
        zakup_df.to_excel(writer, sheet_name='Zakup_Analysis', index=False)
        vip_df.to_excel(writer, sheet_name='VIP_Analysis', index=False)
        
        # Получаем workbook
        workbook = writer.book
        
        # Применяем форматирование к каждому листу
        for sheet_name, df in [('Cash_Analysis', cash_df), ('F17_Analysis', f17_df), 
                              ('Zakup_Analysis', zakup_df), ('VIP_Analysis', vip_df)]:
            worksheet = workbook[sheet_name]
            worksheet.freeze_panes = 'A2'
            
            # Применяем соответствующее условное форматирование
            try:
                cols_map = {name: i + 1 for i, name in enumerate(df.columns)}
                last_row = df.shape[0] + 1
                
                # Общее форматирование для всех листов
                # раньше было: if "Delta" in cols_map: ...
                delta_key = "Delta" if "Delta" in cols_map else ("Delta_F17" if "Delta_F17" in cols_map else None)
                if delta_key:
                    delta_col = get_column_letter(cols_map[delta_key])
                    delta_range = f"{delta_col}2:{delta_col}{last_row}"
                    worksheet.conditional_formatting.add(
                        delta_range,
                        ColorScaleRule(
                            start_type="num", start_value=-1, start_color="FF6B6B",
                            mid_type="num", mid_value=0, mid_color="D3D3D3",
                            end_type="num", end_value=1, end_color="90EE90"
                        )
                    )
                
                # Margin_to_Cost_% — 3-цветная шкала (красный→жёлтый→зелёный)
                if "Margin_to_Cost_%" in cols_map:
                    margin_col = get_column_letter(cols_map["Margin_to_Cost_%"])
                    margin_range = f"{margin_col}2:{margin_col}{last_row}"
                    
                    worksheet.conditional_formatting.add(
                        margin_range,
                        ColorScaleRule(
                            start_type='min', start_color='FFF8696B',
                            mid_type='percentile', mid_value=50, mid_color='FFFFEB84',
                            end_type='max', end_color='FF63BE7B'
                        )
                    )
                
                # Специфичное форматирование для каждого типа анализа
                if sheet_name == 'Cash_Analysis':
                    # Градиент для Delta_Cost_to_CashNew_%
                    if "Delta_Cost_to_CashNew_%" in cols_map:
                        col_letter = get_column_letter(cols_map["Delta_Cost_to_CashNew_%"])
                        rng = f"{col_letter}2:{col_letter}{last_row}"
                        worksheet.conditional_formatting.add(
                            rng,
                            ColorScaleRule(
                                start_type='min', start_color='FFF8696B',
                                mid_type='percentile', mid_value=50, mid_color='FFFFEB84',
                                end_type='max', end_color='FF63BE7B'
                            )
                        )
                    
                    # Cash_new: синий если изменилась цена
                    if "Cash_new" in cols_map and "Cash_old" in cols_map:
                        cash_new_col = get_column_letter(cols_map["Cash_new"])
                        cash_old_col = get_column_letter(cols_map["Cash_old"])
                        cash_new_range = f"{cash_new_col}2:{cash_new_col}{last_row}"
                        
                        worksheet.conditional_formatting.add(
                            cash_new_range,
                            FormulaRule(
                                formula=[f"AND({cash_new_col}2<>{cash_old_col}2,{cash_new_col}2<>\"\")"],
                                fill=PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                            )
                        )
                    
                    empty_cols = ["Cost", "Cash_old", "Cash_new"]
                
                elif sheet_name == 'F17_Analysis':
                    # F17_new: синий если изменилась цена
                    if "F17_new" in cols_map and "F17_old" in cols_map:
                        f17_new_col = get_column_letter(cols_map["F17_new"])
                        f17_old_col = get_column_letter(cols_map["F17_old"])
                        f17_new_range = f"{f17_new_col}2:{f17_new_col}{last_row}"
                        
                        worksheet.conditional_formatting.add(
                            f17_new_range,
                            FormulaRule(
                                formula=[f"AND({f17_new_col}2<>{f17_old_col}2,{f17_new_col}2<>\"\")"],
                                fill=PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                            )
                        )
                    
                    empty_cols = ["Cost", "Cash", "F17_old", "F17_new"]
                
                elif sheet_name == 'Zakup_Analysis':
                    # Zakup_new: синий если изменилась цена
                    if "Zakup_new" in cols_map and "Zakup_old" in cols_map:
                        zakup_new_col = get_column_letter(cols_map["Zakup_new"])
                        zakup_old_col = get_column_letter(cols_map["Zakup_old"])
                        zakup_new_range = f"{zakup_new_col}2:{zakup_new_col}{last_row}"
                        
                        worksheet.conditional_formatting.add(
                            zakup_new_range,
                            FormulaRule(
                                formula=[f"AND({zakup_new_col}2<>{zakup_old_col}2,{zakup_new_col}2<>\"\")"],
                                fill=PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                            )
                        )
                    
                    empty_cols = ["Cost", "Zakup_old", "Zakup_new"]
                
                elif sheet_name == 'VIP_Analysis':
                    # New_VIP: синий если изменилась цена
                    if "New_VIP" in cols_map and "Old_VIP" in cols_map:
                        new_vip_col = get_column_letter(cols_map["New_VIP"])
                        old_vip_col = get_column_letter(cols_map["Old_VIP"])
                        new_vip_range = f"{new_vip_col}2:{new_vip_col}{last_row}"
                        
                        worksheet.conditional_formatting.add(
                            new_vip_range,
                            FormulaRule(
                                formula=[f"AND({new_vip_col}2<>{old_vip_col}2,{new_vip_col}2<>\"\")"],
                                fill=PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                            )
                        )
                    
                    empty_cols = ["Cost", "Old_VIP", "New_VIP"]
                    fill_gray = PatternFill(fill_type="solid", fgColor="FFD9D9D9")
                    
                    # Пустые значения: светло-серый (только для VIP)
                    for col_name in empty_cols:
                        if col_name in cols_map:
                            col_letter = get_column_letter(cols_map[col_name])
                            col_range = f"{col_letter}2:{col_letter}{last_row}"
                            
                            worksheet.conditional_formatting.add(
                                col_range,
                                FormulaRule(
                                    formula=[f'OR(ISBLANK({col_letter}2),LEN(TRIM({col_letter}2))=0,{col_letter}2="")'],
                                    fill=fill_gray,
                                    stopIfTrue=False
                                )
                            )
                    
                    # Выделить всю строку красным, если есть остаток (Stock>0), но New_VIP пустой
                    if "Stock" in cols_map and "New_VIP" in cols_map:
                        stock_letter = get_column_letter(cols_map["Stock"])
                        newvip_letter = get_column_letter(cols_map["New_VIP"])
                        last_col_letter = get_column_letter(len(df.columns))
                        
                        formula = f'AND(${stock_letter}2>0, LEN(TRIM(${newvip_letter}2&""))=0)'
                        row_range = f"A2:{last_col_letter}{last_row}"
                        
                        worksheet.conditional_formatting.add(
                            row_range,
                            FormulaRule(
                                formula=[formula],
                                fill=PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                            )
                        )
                        
            except Exception as e:
                st.warning(f"Ошибка условного форматирования для {sheet_name}: {e}")
    
    output.seek(0)
    return output

# ========= UI =========
st.title("Объединенный анализ цен: Cash, F17, Zakup, VIP")

st.write(
    "**Объединенный анализ всех цен** - Cash, F17, Zakup, VIP в одном Excel файле"
)
st.write("- Загрузите основной .xlsx с рынком (жёлтые флаги)")
st.write("- Результат: Excel файл с 4 листами (Cash, F17, Zakup, VIP)")

# Загрузка файлов
upl_file = st.file_uploader("Основной .xlsx (рынок)", type=["xlsx"], key="market_file")

st.write("---")

if st.button("Запустить все анализы"):
    if upl_file is not None:
        try:
            # Показываем прогресс
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            # Загружаем данные с флагами
            status_text.text("Загрузка данных с флагами...")
            df_with_flags = add_yellow_flags_to_df(upl_file)
            progress_bar.progress(10)
            
            progress_bar.progress(20)
            
            # Запускаем все анализы
            status_text.text("Запуск Cash анализа...")
            cash_df = analyze_cash_df(df_with_flags)
            progress_bar.progress(40)
            
            status_text.text("Запуск F17 анализа...")
            f17_df = analyze_f17_df(df_with_flags)
            progress_bar.progress(60)
            
            status_text.text("Запуск Zakup анализа...")
            zakup_df = analyze_zakup_df(df_with_flags)
            progress_bar.progress(80)
            
            status_text.text("Запуск VIP анализа...")
            vip_df = analyze_vip_df(df_with_flags)
            progress_bar.progress(90)
            
            # Показываем результаты
            status_text.text("Отображение результатов...")
            
            # Создаем табы для отображения результатов
            tab1, tab2, tab3, tab4 = st.tabs(["Cash", "F17", "Zakup", "VIP"])
            
            with tab1:
                st.dataframe(cash_df, height=400)
                st.write(f"**Cash анализ**: {len(cash_df)} позиций")
            
            with tab2:
                st.dataframe(f17_df, height=400)
                st.write(f"**F17 анализ**: {len(f17_df)} позиций")
            
            with tab3:
                st.dataframe(zakup_df, height=400)
                st.write(f"**Zakup анализ**: {len(zakup_df)} позиций")
            
            with tab4:
                st.write("**VIP-анализ Excel: жёсткие правила**")
                st.write(
                    "- Жёлтые/перечёркнутые берём из форматирования.\n"
                    "- Old_VIP учитываем только если ячейка VIP жёлтая и НЕ перечёркнута.\n"
                    "- Cost участвует только если ячейка Cost жёлтая (>0). При Stock=0 Cost снимаем.\n"
                    "- Остаток: суммируем все склады, КРОМЕ последнего (виртуал); Virtual_Stock — последний сегмент.\n"
                    "- При Virtual_Stock=35 расчёт VIP не делаем. При Stock=0 и Virtual=0 — тоже пропуск.\n"
                    "- Market_suppliers_all/count — все рыночные поставщики по позиции.\n"
                )
                
                highlight_mode = st.selectbox(
                    "Подсветка New_VIP:",
                    ["По флагу Change == 'Y'", "По неравенству New_VIP != Old_VIP"],
                    index=0,
                    key="newvip_highlight_mode"
                )
                
                total = len(vip_df)
                changed = int((vip_df["Change"] == "Y").sum()) if "Change" in vip_df.columns else 0
                no_new = int(vip_df["New_VIP"].isna().sum()) if "New_VIP" in vip_df.columns else 0
                forced = int((vip_df["Cost_Forced_NoCost"] == "Y").sum()) if "Cost_Forced_NoCost" in vip_df.columns else 0
                st.caption(f"Найдено товаров: {total} | Изменений: {changed} | Без нового VIP: {no_new} | Cost снят принудительно: {forced}")
                
                st.dataframe(vip_df, height=620)
                
                # ===== VIP Excel =====
                vip_output = BytesIO()
                with pd.ExcelWriter(vip_output, engine="openpyxl") as writer:
                    vip_df.to_excel(writer, index=False, sheet_name="VIP_result")
                    ws = writer.sheets["VIP_result"]
                    ws.freeze_panes = "A2"

                    # автоширина
                    from openpyxl.utils import get_column_letter
                    for cidx, col in enumerate(vip_df.columns, start=1):
                        series = vip_df[col].astype(str)
                        max_len = max([len(col)] + [len(s) for s in series]) if len(series) else len(col)
                        ws.column_dimensions[get_column_letter(cidx)].width = min(max_len + 2, 60)

                    # заливки
                    from openpyxl.styles import PatternFill
                    from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
                    fill_gray  = PatternFill(fill_type="solid", fgColor="FFD9D9D9")
                    fill_blue  = PatternFill(fill_type="solid", fgColor="FFCFE2F3")
                    cols_map = {name: i + 1 for i, name in enumerate(vip_df.columns)}
                    last_row = vip_df.shape[0] + 1

                    def rng(col):
                        if col not in cols_map: return None, None
                        L = get_column_letter(cols_map[col])
                        return f"{L}2:{L}{last_row}", L

                    # пустые Cost/Old_VIP/New_VIP → серым
                    for cname in ["Cost","Old_VIP","New_VIP"]:
                        arng, L = rng(cname)
                        if arng:
                            ws.conditional_formatting.add(
                                arng,
                                FormulaRule(
                                    formula=[f'OR(ISBLANK({L}2),LEN(TRIM({L}2))=0,{L}2="")'],
                                    fill=fill_gray,
                                    stopIfTrue=False
                                )
                            )
                            # явная подстраховка
                            for r in range(2, last_row + 1):
                                cell = ws[f"{L}{r}"]
                                if cell.value is None or (isinstance(cell.value, str) and str(cell.value).strip()==""):
                                    cell.fill = fill_gray

                    # Delta: -1 → 0 → 1
                    drng, _ = rng("Delta")
                    if drng:
                        ws.conditional_formatting.add(
                            drng,
                            ColorScaleRule(
                                start_type='num', start_value=-1, start_color='FFF4CCCC',
                                mid_type='num', mid_value=0,  mid_color='FFD9D9D9',
                                end_type='num', end_value=1,  end_color='FF93C47D'
                            )
                        )

                    # Margin_to_Cost_%: красн→жёлт→зел
                    mrng, _ = rng("Margin_to_Cost_%")
                    if mrng:
                        ws.conditional_formatting.add(
                            mrng,
                            ColorScaleRule(
                                start_type='min', start_color='FFF8696B',
                                mid_type='percentile', mid_value=50, mid_color='FFFFEB84',
                                end_type='max', end_color='FF63BE7B'
                            )
                        )

                    # Красим строку красным, если Stock>0, а New_VIP пустой
                    if "Stock" in cols_map and "New_VIP" in cols_map:
                        stockL = get_column_letter(cols_map["Stock"])
                        newL   = get_column_letter(cols_map["New_VIP"])
                        lastL  = get_column_letter(len(vip_df.columns))
                        row_range = f"A2:{lastL}{last_row}"
                        ws.conditional_formatting.add(
                            row_range,
                            FormulaRule(
                                formula=[f'AND(${stockL}2>0, LEN(TRIM(${newL}2&""))=0)'],
                                fill=PatternFill(fill_type="solid", fgColor="FFF4CCCC"),
                                stopIfTrue=False
                            )
                        )

                    # Подсветка New_VIP (режимы)
                    nvrng, newL = rng("New_VIP")
                    if nvrng:
                        if highlight_mode == "По флагу Change == 'Y'" and "Change" in cols_map:
                            chL = get_column_letter(cols_map["Change"])
                            ws.conditional_formatting.add(
                                nvrng,
                                FormulaRule(formula=[f'${chL}2="Y"'], fill=fill_blue, stopIfTrue=False)
                            )
                        elif highlight_mode == "По неравенству New_VIP != Old_VIP" and "Old_VIP" in cols_map:
                            oldL = get_column_letter(cols_map["Old_VIP"])
                            ws.conditional_formatting.add(
                                nvrng,
                                FormulaRule(
                                    formula=[f'AND(LEN(TRIM(${oldL}2))>0,LEN(TRIM(${newL}2))>0,${newL}2<>${oldL}2)'],
                                    fill=fill_blue, stopIfTrue=False
                                )
                            )

                vip_output.seek(0)
                st.download_button(
                    label="Скачать VIP Excel",
                    data=vip_output.getvalue(),
                    file_name="vip_analysis.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            
            # Создаем объединенный Excel файл
            status_text.text("Создание Excel файла...")
            excel_output = export_all_analyses_to_excel(cash_df, f17_df, zakup_df, vip_df)
            progress_bar.progress(100)
            
            # Кнопка скачивания объединенного Excel
            st.download_button(
                label="📊 Скачать объединенный Excel (все анализы)",
                data=excel_output.getvalue(),
                file_name="unified_price_analysis.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            
            status_text.text("✅ Все анализы завершены!")
            st.success("🎉 Все анализы завершены! Excel файл с 4 листами готов к скачиванию.")
            
        except Exception as e:
            st.error(f"Ошибка: {e}")
            import traceback
            st.error(f"Детали ошибки: {traceback.format_exc()}")
    else:
        st.warning("Пожалуйста, загрузите основной файл с рынком.")

# ========= SELFCHECK PRICES =========
def _selfcheck_prices():
    """Дымовой чек Cash/VIP - валит приложение при критических изменениях"""
    # 1) Cash таргетит p2 и уважает cost+1.5%
    result, reason, rank = cash_choose_price_p2([100, 120, 130], 110, None)
    assert result >= 111.5, f"Cash должен соблюдать маржу 1.5%, got {result}"
    
    result, reason, rank = cash_choose_price_p2([100, 120, 130], None, 140)
    assert result == 140, f"Cash должен использовать cash_old при отсутствии cost, got {result}"
    
    # 2) trim_by_gap должен отбрасывать выбросы
    prices = [100.0, 120.0, 200.0]  # 200 - выброс (gap > 40%)
    cleaned, reason = trim_by_gap(prices)
    assert len(cleaned) == 2, f"trim_by_gap должен отбросить выброс, got {cleaned}"
    assert reason == "DROP_MAX_BY_GAP40", f"Неправильная причина: {reason}"
    
    # 3) VIP: базовые проверки (пока простые, можно расширить)
    # VIP функции используют замороженные константы через параметры
    
    print("✅ SELFCHECK PRICES PASSED")

# Запускаем чек при импорте (под флагом DEV)
DEV_MODE = True  # можно выключить в продакшене
if __name__ == "__main__" and DEV_MODE:
    try:
        _selfcheck_prices()
    except Exception as e:
        print(f"❌ SELFCHECK PRICES FAILED: {e}")
        raise

